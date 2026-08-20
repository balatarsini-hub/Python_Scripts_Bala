# %% [markdown]
# # SG CPD COMMENTARY SCRIPT
# 

# %%
import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta

dir = os.getcwd()

# Get the current date
current_date = datetime.now()

# Calculate the previous month
previous_month = current_date - relativedelta(months=1)

# Create folder structure for the month inside the subfolder
month_subfolder_name = previous_month.strftime("%Y-%m")


#Specify folder for commentary
generated_report_folder = f'{dir}/Full Data'
commentary_template_folder = f'{dir}/Template/Commentary Template'
output_report_folder = f'{dir}/Generated Report/Commentary Report'
generated_report_folder

# %% [markdown]
# Load the file
# 

# %%
# Choosing file to use
desired_generated_report_file = r"C:\Users\balatarsini_avinitya\Downloads\Topline Report - July\Full Data\SG CPD&LDB Female&Male Skincare Topline_Final 2025-12-29 Corn (1)_CLEAN_full.xlsx"

# Construct the full path to the specified generated data file
generated_report_path = os.path.join(generated_report_folder, desired_generated_report_file)

# Load all sheets of the generated data into a dictionary
generated_report = pd.read_excel(generated_report_path, sheet_name=None)

# Load the empty template from 'Commentary Template' with multiple sheets
template_path_commentary = os.path.join(commentary_template_folder, "SG Skincare CPD Commentary Template.xlsx")

# %% [markdown]
# # Important Sheets for Commentary
# 

# %%
#Male
sheet_data  = generated_report['1-Table-1']
sheet_data1 = generated_report['2-Table-1']
sheet_data2 = generated_report['3-Table-1']
sheet_data3 = generated_report['4-Table-1']

#Female
sheet_data4 = generated_report['7-Table-1']
sheet_data5 = generated_report['8-Table-1']
sheet_data6 = generated_report['9-Table-1']
sheet_data7 = generated_report['10-Table-1']

# %% [markdown]
# # MEN CPD

# %% [markdown]
# CPD Summary Table
# 

# %%
cpdsummary_df = sheet_data.iloc[13:49, 0:38] #change this based on data
cpdsummary_df1 = sheet_data2.iloc[13:49, 0:38]
cpdsummary_df4 = sheet_data1.iloc[:,0:38]

# Retrieve and format the month and year
month_year = cpdsummary_df4.columns[35]
month_year_formatted = f"{month_year[:3]}'{month_year[-2:]} vs. {month_year[:3]}'{int(month_year[-2:])-1:02d}"

# Retrieve and format the market sales value
sales_value = cpdsummary_df4.iloc[1, 35]
if sales_value >= 1e6:
    sales_value_formatted = f"{sales_value / 1e6:.1f} mil"
else:
    sales_value_formatted = f"{sales_value / 1e3:.1f} thousand"

# Retrieve and format the evolution percentage
evo_percentage = cpdsummary_df4.iloc[1, 36]
evo_percentage_formatted = f"{evo_percentage:.1f}%"

# Construct the final sentence
final_sentence = f"{month_year_formatted}: Market sales at {sales_value_formatted} with {evo_percentage_formatted} evo."

# Create a DataFrame with the final_sentence
final_sentence_df1 = pd.DataFrame([[final_sentence]])

# list of brands to filter
brands_to_filter = ['TOTAL CPD', 'GARNIER', "LOREAL DERMO EXPERTISE"]

# find the correct column name containing brand information
brand_column = cpdsummary_df.columns[0]
brand_column1 = cpdsummary_df1.columns[0]

# filter the DataFrame based on the specified brands
filtered_rows = cpdsummary_df[cpdsummary_df[brand_column].isin(brands_to_filter)]
filtered_rows1 = cpdsummary_df1[cpdsummary_df1[brand_column1].isin(brands_to_filter)]

# Value (this month)
Market_Value = sheet_data1.iloc[1:2, 36:37]/100 #correct
Ms_Value = filtered_rows.iloc[:,35:36]/100
Evo_Value = filtered_rows.iloc[:,36:37]/100

# Unit (this month)
Market_Unit =sheet_data3.iloc[1:2, 36:37]/100
Ms_Unit = filtered_rows1.iloc[:,35:36] /100
Evo_Unit = filtered_rows1.iloc[:,36:37]/100

# Value (Last month)
Market_Value1 = sheet_data1.iloc[1:2, 34:35]/100
Ms_Value1 = filtered_rows.iloc[:,33:34]/100
Evo_Value1 = filtered_rows.iloc[:,34:35]/100

# Unit (Last month)
Market_Unit1 =sheet_data3.iloc[1:2, 34:35]/100
Ms_Unit1 = filtered_rows1.iloc[:,33:34] /100
Evo_Unit1 = filtered_rows1.iloc[:,34:35]/100

# Value (Last 2 month)
Market_Value2 = sheet_data1.iloc[1:2, 32:33]/100
Ms_Value2 = filtered_rows.iloc[:,31:32]/100
Evo_Value2 = filtered_rows.iloc[:,32:33]/100

# Unit (Last 2 month)
Market_Unit2 =sheet_data3.iloc[1:2, 32:33]/100
Ms_Unit2 = filtered_rows1.iloc[:,31:32] /100
Evo_Unit2 = filtered_rows1.iloc[:,32:33]/100

# %% [markdown]
# For Dates on Table

# %%
import re
from datetime import datetime, timedelta
import calendar

# Function to extract month and year
def extract_month_year(month_year):
    match = re.match(r"(\w{3})\s*(\d{2})", month_year)
    if match:
        month_str, year_str = match.groups()
        month = datetime.strptime(month_str, "%b").month
        year = int('20' + year_str)  # Adjust if year needs to be in 1900s or 2000s
        return month, year
    return None, None

# Function to get month name with year
def get_month_name_with_year(month, year):
    month_str = calendar.month_abbr[month]
    year_str = str(year)[2:]  # Get last 2 digits of the year
    return f"{month_str}'{year_str}"

# Function to get the month before a given month and year
def get_month_before(month, year, num_months):
    # Handle month subtraction with year transition
    for _ in range(num_months):
        if month == 1:
            month = 12
            year -= 1
        else:
            month -= 1
    return month, year

# Retrieve and clean column names
current_month = cpdsummary_df.columns[35]
month_1_before = cpdsummary_df.columns[34]

# Extract month and year
current_month_num, current_year = extract_month_year(current_month)
month_1_before_num, month_1_before_year = extract_month_year(month_1_before)

# Debug output
print(f"Current month column: {current_month}, Month number: {current_month_num}, Year: {current_year}")
print(f"Month 1 before column: {month_1_before}, Month number: {month_1_before_num}, Year: {month_1_before_year}")

# Check if extracted values are valid
if current_month_num is None or current_year is None:
    raise ValueError("Current month or year could not be extracted.")
if month_1_before_num is None or month_1_before_year is None:
    raise ValueError("Month 1 before could not be extracted.")

# Get the month before
month_2_before_num, month_2_before_year = get_month_before(current_month_num, current_year, 2)

date1 = get_month_name_with_year(current_month_num, current_year)
date2 = get_month_name_with_year(month_1_before_num, month_1_before_year)
date3 = get_month_name_with_year(month_2_before_num, month_2_before_year)


# %% [markdown]
# Market Summary Table

# %% MALE CPD
#Function
combine = pd.concat([sheet_data, sheet_data2], axis=1)
functionsummary_df = combine.iloc[2:8,0:74] #change this based on data

column_names_function = ['Function','2022','2022 Evo', '2023', '2023 Evo', 'YTD-1','YTD-1 Evo','YTD' ,'YTD Evo', 'MAT-1', 'MAT-1 Evo','MAT','MAT Evo','Month1','Month1 Evo','Month2', 'Month2 Evo', 
                        'Month3','Month3 Evo','Month4','Month4 Evo', 'Month5','Month5 Evo', 'Month6','Month6 Evo','Month7','Month7 Evo','Month8','Month8 Evo','Month9','Month9 Evo','Month10','Month10 Evo',
                        'Month11','Month11 Evo','Month12','Month12 Evo','Function1','2022 Unit','2022 Unit Evo', '2023Unit','2023 Unit Evo','YTD-1 Unit','YTD-1 Unit Evo','YTD Unit' ,'YTD Unit Evo','MAT-1 Unit', 
                        'MAT-1 Unit Evo','MAT Unit','MAT Unit Evo','Month1 Unit','Month1 Unit Evo','Month2 Unit' ,'Month2 Unit Evo','Month3 Unit','Month3 Unit Evo','Month4 Unit','Month4 Unit Evo', 'Month5 Unit',
                        'Month5 Unit Evo', 'Month6 Unit','Month6 Unit Evo','Month7 Unit','Month7 Unit Evo','Month8 Unit','Month8 Unit Evo','Month9 Unit','Month9 Unit Evo','Month10 Unit','Month10 Unit Evo',
                        'Month11 Unit','Month11 Unit Evo','Month12 Unit','Month12 Unit Evo']
functionsummary_df.columns = column_names_function

replacements = {'OIL CONTROL': 'Oil Control',
                'AGING': 'Aging',
                'BRIGHTENING': 'Brightening',
                'HYDRATION + BASIC': 'Hydration + Basic'}

# %%
#Function
function_df_asc = functionsummary_df[~functionsummary_df['Function'].isin(['Hydration','Basic'])][['Function', 'Month12 Evo', 'Month12 Unit Evo']].sort_values(by='Month12 Evo', ascending=False)
function_df_asc['Month12 Evo'] /= 100
function_df_asc['Month12 Unit Evo'] /= 100
function_df_asc['Function'] = function_df_asc['Function'].replace(replacements)

# %%
#Format
formatsummary_df = combine.iloc[8:13,0:74]
formatsummary_df.columns = column_names_function

replacements = {'CLEANSER': 'Cleanser',
                'MOISTURISER': 'Moisturiser',
                'FACE MASK': 'Face Mask',
                'MAKEUP REMOVER MICELLAR WATER':'Make Up Remover',
                'SCRUB':'Scrub',
                'SERUM':'Serum'}

#this month
format_df_asc = formatsummary_df[['Function', 'Month12 Evo', 'Month12 Unit Evo']].sort_values(by='Month12 Evo', ascending=False)
format_df_asc['Month12 Evo'] /= 100
format_df_asc['Month12 Unit Evo'] /= 100
format_df_asc['Function'] = format_df_asc['Function'].replace(replacements)

print(format_df_asc)

# %% MALE CPD
#Brand
brandsummary_df = combine.iloc[14:50,0:74] #need to change this based on the data
brandsummary_df.columns = column_names_function

# brand names
brand_names = ['GARNIER',"L'OREAL DERMO EXPERTISE",'Nivea','Gatsby',"Men's Biore",'Shokobutsu','Eversoft','Nano White','Sukin','Other Brands','Skintific', 'Nanowhite']

def identify_brand(brand):
    if brand in brand_names:
        return 'brand'
    else:
        return 'aux'

brandsummary_df['Brand_Type'] = brandsummary_df['Function'].apply(identify_brand)

replacements = {'GARNIER': 'Garnier',
                "L'OREAL DERMO EXPERTISE": "L'oreal Paris"}

# %%
# Select data for this month
top_brands_df = brandsummary_df[brandsummary_df['Brand_Type'] == 'brand'][['Function', 'Month12 Evo', 'Month12 Unit Evo']].sort_values(by = 'Month12 Evo', ascending = False).head(3)
top_brands_df['Month12 Evo'] /= 100
top_brands_df['Month12 Unit Evo'] /= 100
top_brands_df['Function'] = top_brands_df['Function'].replace(replacements)

top_brands_df

# %% [markdown]
# For YTD sentences at the bottom

# %% MALE CPD
def ahead_or_behind(growth_value, value):
    c, m = growth_value, value

    if m < 0 and c < 0:
        # both negative
        if c < m:      # company more negative
            return "behind"    # (1)
        elif c > m:    # company less negative
            return "ahead"     # (2)
        else:
            return "inline"

    if m < 0 and c >= 0:
        return "ahead"          # (3)

    if m > 0 and c > 0:
        if c > m:
            return "ahead"      # (4)
        elif c < m:
            return "behind"     # (5)
        else:
            return "inline"

    if m > 0 and c <= 0:
        return "behind"         # (6)

    # market == 0 fallback
    if m == 0:
        if c > 0:  return "ahead"
        if c < 0:  return "behind"
        return "inline"

# %% MALE CPD
# Assuming `date1` is already defined and contains the formatted month and year
sentence0 = f"YTD {date1}"
print(sentence0)

# Access the value from row 3, column I (8th column, as 'A' is 0)
value = sheet_data.iloc[1, 8]  # Use iloc to access row 3 (index 1) and column I (index 8)

# Determine if the market grew or declined
if value is not None:
    rounded_value = round(value, 1)  # Round the value to one decimal place
    if rounded_value > 0:
        sentence1 = f"1. Market grew +{rounded_value}% vs LY"
    else:
        sentence1 = f"1. Market declined {rounded_value}% vs LY"
else:
    sentence1 = "1. Value in cell 3I is not available."

print(sentence1)

# Access the values from the required cells
growth_value = sheet_data.iloc[13, 8]  # Value from cell 14I (index 13, column index 8)
market_share_current = sheet_data.iloc[13, 7]  # Value from cell 14H (index 13, column index 7)
market_share_ly = sheet_data.iloc[13, 5]  # Value from cell 14F (index 13, column index 5)

# Calculate rounded growth value and ratio for "ahead/behind" calculation
rounded_growth_value = round(growth_value, 1)
ratio_ahead_behind = round(abs(growth_value / value), 1) if value != 0 else 0.0
ahead_behind = ahead_or_behind(growth_value, value)


# Calculate the market share difference
market_share_difference = round(market_share_current - market_share_ly, 1)

# Round all values to one decimal place for consistent formatting
market_share_current_rounded = round(market_share_current, 1)

# Generate the growth sentence
if rounded_growth_value > 0:
    growth_sentence = (
        f"2. Loreal CPD grew +{rounded_growth_value:.1f}%, "
        f"{ratio_ahead_behind:.1f}x {ahead_behind} of market with "
        f"{market_share_current_rounded:.1f}%MS ({market_share_difference:+.1f} pp MS vs LY)"
    )
else:
    growth_sentence = (
        f"2. Loreal CPD declined {rounded_growth_value:.1f}%, "
        f"{ratio_ahead_behind:.1f}x {ahead_behind} of market with "
        f"{market_share_current_rounded:.1f}%MS ({market_share_difference:+.1f} pp MS vs LY)"
    )

print(growth_sentence)

# List of brands to include in the final sentence
brands_of_interest = [
    "GARNIER", "L'Oreal Paris", "Nivea", "Gatsby", "Men's Biore", "Shokobutsu", "Eversoft", 
    "Nanowhite", "Sukin", "Other Brands"
]

# Access the DataFrame for the specified rows and columns, filtering by the brands of interest
brands = sheet_data.iloc[14:46, 0]  # Brands from row 15 to row 46 (index 14 to 45)
column_h_values = sheet_data.iloc[14:46, 7]  # Values from column H
column_f_values = sheet_data.iloc[14:46, 5]  # Values from column F

# Initialize variables to keep track of the best result
max_difference = float('-inf')
best_brand = None

# Loop through the brands to find the one with the highest difference
for i, brand in enumerate(brands):
    if brand == "LOREAL DERMO EXPERTISE":
        brand = "L'Oreal Paris"  # Convert to L'Oreal Paris
    if brand in brands_of_interest:  # Check if the brand is in the list of brands of interest
        difference = column_h_values.iloc[i] - column_f_values.iloc[i]  # Calculate difference
        if difference > max_difference:
            max_difference = difference
            best_brand = brand

# Generate the sentence with the highest difference
if best_brand is not None:
    rounded_difference = round(max_difference, 1)  # Round the difference to one decimal place
    sentence = f"3. Share gain led by {best_brand} (+{rounded_difference:.1f} pp MS vs LY)"
else:
    sentence = "3. No share gain data available."

print(sentence)


# %% [markdown]
# # FEMALE CPD

# %% [markdown]
# CPD Summary Table

# %% FEMALE CPD
cpdsummary_df2 = sheet_data4.iloc[19:82, 0:38]
cpdsummary_df3 = sheet_data6.iloc[19:82, 0:38]
cpdsummary_df5 = sheet_data5.iloc[:, 0:38]

# Retrieve and format the month and year
month_year = cpdsummary_df5.columns[35]
month_year_formatted = f"{month_year[:3]}'{month_year[-2:]} vs. {month_year[:3]}'{int(month_year[-2:])-1:02d}"

# Retrieve and format the market sales value
sales_value = cpdsummary_df5.iloc[1, 35]
if sales_value >= 1e6:
    sales_value_formatted = f"{sales_value / 1e6:.1f} mil"
else:
    sales_value_formatted = f"{sales_value / 1e3:.1f} thousand"

# Retrieve and format the evolution percentage
evo_percentage = cpdsummary_df5.iloc[1, 36]
evo_percentage_formatted = f"{evo_percentage:.1f}%"

# Construct the final sentence
final_sentence = f"{month_year_formatted}: Market sales at {sales_value_formatted} with {evo_percentage_formatted} evo."

# Create a DataFrame with the final_sentence
final_sentence_df2 = pd.DataFrame([[final_sentence]])

# list of brands to filter
brands_to_filter1 = ['TOTAL CPD', 'GARNIER', "LOREAL DERMO EXPERTISE"]

# find the correct column name containing brand information
brand_column = cpdsummary_df2.columns[0]
brand_column1 = cpdsummary_df3.columns[0]

# filter the DataFrame based on the specified brands
filtered_rows2 = cpdsummary_df2[cpdsummary_df2[brand_column].isin(brands_to_filter1)]
print(filtered_rows2)
filtered_rows3 = cpdsummary_df3[cpdsummary_df3[brand_column1].isin(brands_to_filter1)]

# Value (this month)
Market_Value3 = sheet_data5.iloc[1:2, 36:37]/100
Ms_Value3 = filtered_rows2.iloc[:,35:36]/100
Evo_Value3 = filtered_rows2.iloc[:,36:37]/100
print(Ms_Value3)

# Unit (this month)
Market_Unit3 =sheet_data7.iloc[1:2, 36:37]/100
Ms_Unit3 = filtered_rows3.iloc[:,35:36] /100
Evo_Unit3 = filtered_rows3.iloc[:,36:37]/100

# # Value (Last month)
Market_Value4 = sheet_data5.iloc[1:2, 34:35]/100
Ms_Value4 = filtered_rows2.iloc[:,33:34]/100
Evo_Value4 = filtered_rows2.iloc[:,34:35]/100

# Unit (Last month)
Market_Unit4 =sheet_data7.iloc[1:2, 34:35]/100
Ms_Unit4 = filtered_rows3.iloc[:,33:34] /100
Evo_Unit4 = filtered_rows3.iloc[:,34:35]/100

# Value (Last 2 month)
Market_Value5 = sheet_data5.iloc[1:2, 32:33]/100
Ms_Value5 = filtered_rows2.iloc[:,31:32]/100
Evo_Value5 = filtered_rows2.iloc[:,32:33]/100

# Unit (Last 2 month)
Market_Unit5 =sheet_data7.iloc[1:2, 32:33]/100
Ms_Unit5 = filtered_rows3.iloc[:,31:32] /100
Evo_Unit5 = filtered_rows3.iloc[:,32:33]/100

# %% [markdown]
# Market Summary Table

# %%
#concatenate df 
combine1 = pd.concat([sheet_data4, sheet_data6], axis=1)
functionsummary_df1 = combine1.iloc[4:10,0:74]
print(functionsummary_df1)
functionsummary_df1.columns = column_names_function

replacements = {'WHITEN': 'Female Brightening',
                'HYDRATING': 'Hydrating',
                'BASIC': 'Basic', 
                'ANTI AGING': 'Female Anti-Age',
                'OIL CONTROL': 'Female Oil Control',
                'PURIFYING' : 'Female Purifying'}

#Function
# Selecting data for this month
function_df_asc3 = functionsummary_df1[~functionsummary_df1['Function'].isin(['BASIC HYDRATING'])][['Function', 'Month12 Evo', 'Month12 Unit Evo']].sort_values(by='Month12 Evo', ascending=False)
function_df_asc3['Month12 Evo'] /= 100
function_df_asc3['Month12 Unit Evo'] /= 100
function_df_asc3['Function'] = function_df_asc3['Function'].replace(replacements)
function_df_asc3


# %%
#Format
formatsummary_df1 = combine1.iloc[13:39,0:74]
formatsummary_df1.columns = column_names_function

replacements = {'CLEANSER': 'Cleanser',
                'MOISTURISER': 'Moisturiser',
                'SCRUB': 'Scrub',
                'MOISTURIZER ESSENCE':'Serum',
                'MASK': 'Face Mask',
                'TONER': 'Toner',
                'MAKE UP REMOVER': 'Make Up Remover'}

#select data this month
format_df_asc3 = formatsummary_df1[formatsummary_df1['Function'].isin(['CLEANSER', 'MOISTURISER','SCRUB', 'MOISTURIZER ESSENCE', 'MASK', 'TONER', 'MAKE UP REMOVER'])][['Function', 'Month12 Evo', 'Month12 Unit Evo']].sort_values(by = 'Month12 Evo', ascending = False)
format_df_asc3['Month12 Evo'] /= 100
format_df_asc3['Month12 Unit Evo'] /= 100
format_df_asc3['Function'] = format_df_asc3['Function'].replace(replacements)

print(formatsummary_df1)


# %%
#Brand
brandsummary_df1 = combine1.iloc[42:80,0:74]
brandsummary_df1.columns = column_names_function

# brand names
brand_names = [
    "GARNIER",
    "LOREAL DERMO EXPERTISE",
    "MAYBELLINE",
    "HADA LABO",
    "BIO ESSENCE",
    "EVERSOFT",
    "BIORE",
    "OLAY",
    "SKINTIFIC",
    "ST IVES",
    "NUTOX",
    "SAFI",
    "SENKA",
    "Others",
    "EXCLUSIVE + PRIVATE LABELS",
]

def identify_brand(brand):
    if brand in brand_names:
        return 'brand'
    else:
        return 'aux'

# Apply the function to the 'Brand' column
brandsummary_df1['Brand_Type'] = brandsummary_df1['Function'].apply(identify_brand)
print(brandsummary_df1)

replacements = {
    "GARNIER": "Garnier",
    "LOREAL DERMO EXPERTISE": "L'oreal Paris",
    "MAYBELLINE": "Maybelline",
    "HADA LABO": "Hada Labo",
    "BIO ESSENCE": "Bio Essence",
    "EVERSOFT": "Eversoft",
    "BIORE": "Biore",
    "OLAY": "Olay",
    "SKINTIFIC": "Skintific",
    "ST IVES": "St Ives",
    "NUTOX": "Nutox",
    "SAFI": "Safi",
    "SENKA": "Senka",
    "Others": "Others",
    "EXCLUSIVE + PRIVATE LABELS": "Exclusive + Private Labels",
}

brandsummary_df1

# %%
# Select data for this month
top_brands_df3 = brandsummary_df1[brandsummary_df1['Brand_Type'] == 'brand'][['Function', 'Month12 Evo', 'Month12 Unit Evo']].sort_values(by = 'Month12 Evo', ascending = False).head(3)
print(top_brands_df3)
top_brands_df3['Month12 Evo'] /= 100
top_brands_df3['Month12 Unit Evo'] /= 100
top_brands_df3['Function'] = top_brands_df3['Function'].replace(replacements)
top_brands_df3

# %% [markdown]
# For YTD sentences at the bottom

# %%
# Assuming `date1` is already defined and contains the formatted month and year
sentence_0 = f"YTD {date1}"
print(sentence_0)

# Access the value from row 3, column I (8th column, as 'A' is 0)
value = sheet_data5.iloc[1, 8]  # Use iloc to access row 3 (index 1) and column I (index 8)
# Determine if the market grew or declined
if value is not None:
    rounded_value = round(value, 1)  # Round the value to one decimal place
    if rounded_value > 0:
        sentence_1 = f"1. Market grew +{rounded_value}% vs LY"
    else:
        sentence_1 = f"1. Market declined {rounded_value}% vs LY"
else:
    sentence_1 = "1. Value in cell 3I is not available."

print(sentence_1)

# Access the values from the required cells
growth_value = sheet_data4.iloc[37, 8]  # Value from cell 14I (index 13, column index 8)
market_share_current = sheet_data4.iloc[37, 7]  # Value from cell 14H (index 13, column index 7)
market_share_ly = sheet_data4.iloc[37, 5]  # Value from cell 14F (index 13, column index 5)

# Calculate rounded growth value and ratio for "ahead/behind" calculation
rounded_growth_value = round(growth_value, 1)
ratio_ahead_behind = round(abs(growth_value / value), 1) if value != 0 else 0.0
ahead_behind = ahead_or_behind(growth_value, value)

# Calculate the market share difference
market_share_difference = round(market_share_current - market_share_ly, 1)

# Round all values to one decimal place for consistent formatting
market_share_current_rounded = round(market_share_current, 1)

# Generate the growth sentence
if rounded_growth_value > 0:
    growth_sentence1 = (
        f"2. Loreal CPD grew +{rounded_growth_value:.1f}%, "
        f"{ratio_ahead_behind:.1f}x {ahead_behind} of market with "
        f"{market_share_current_rounded:.1f}%MS ({market_share_difference:+.1f} pp MS vs LY)"
    )
else:
    growth_sentence1 = (
        f"2. Loreal CPD declined {rounded_growth_value:.1f}%, "
        f"{ratio_ahead_behind:.1f}x {ahead_behind} of market with "
        f"{market_share_current_rounded:.1f}%MS ({market_share_difference:+.1f} pp MS vs LY)"
    )

print(growth_sentence1)

# List of brands to include in the final sentence
brands_of_interest = [
    "GARNIER", "L'OREAL PARIS", "MAYBELLINE", "HADO LABO", "BIO ESSENCE","Eversoft", 
    "BIORE", "OLAY", "SKINTIFIC","SIMPLE","ST IVES","HIMALAYA","NUTOX","SAFI","SENKA","Others","EXCLUSIVE + PRIVATE LABELS"
]

# Access the DataFrame for the specified rows and columns, filtering by the brands of interest
brands = sheet_data4.iloc[19:88, 0]  # Brands from row 15 to row 46 (index 14 to 45)
column_h_values = sheet_data4.iloc[19:88, 7]  # Values from column H
column_f_values = sheet_data4.iloc[19:88, 5]  # Values from column F

# Initialize variables to keep track of the best result
max_difference = float('-inf')
best_brand = None

# Loop through the brands to find the one with the highest difference
for i, brand in enumerate(brands):
    if brand == "LOREAL DERMO EXPERTISE":
        brand = "L'Oreal Paris"  # Convert to L'Oreal Paris
    if brand in brands_of_interest:  # Check if the brand is in the list of brands of interest
        difference = column_h_values.iloc[i] - column_f_values.iloc[i]  # Calculate difference
        if difference > max_difference:
            max_difference = difference
            best_brand = brand

# Generate the sentence with the highest difference
if best_brand is not None:
    rounded_difference = round(max_difference, 1)  # Round the difference to one decimal place
    sentence_2 = f"3. Share gain led by {best_brand} (+{rounded_difference:.1f} pp MS vs LY)"
else:
    sentence_2 = "3. No share gain data available."

print(sentence_2)



# %% [markdown]
# Load Data into Template

# %%
# Load the template workbook
wb = load_workbook(template_path_commentary)

# Function to load selected data into the template sheet
def load_data_into_template(data, sheet, start_row, start_column):
    for i, row in enumerate(data.values, start=start_row):
        for j, value in enumerate(row, start=start_column):
            sheet.cell(row=i, column=j, value=value)
            
# Men CPD Summary Table
#this month
load_data_into_template(Market_Value, wb["Men CPD"], 8, 6)
load_data_into_template(Market_Unit, wb["Men CPD"], 8, 11)
load_data_into_template(Ms_Value, wb["Men CPD"], 9, 5)
load_data_into_template(Evo_Value, wb["Men CPD"], 9, 6)
load_data_into_template(Ms_Unit, wb["Men CPD"], 9, 10)
load_data_into_template(Evo_Unit, wb["Men CPD"], 9, 11)

#last month
load_data_into_template(Ms_Value1, wb["Men CPD"], 9, 4)
load_data_into_template(Ms_Unit1, wb["Men CPD"], 9, 9)

#last 2 month
load_data_into_template(Ms_Value2, wb["Men CPD"], 9, 3)
load_data_into_template(Ms_Unit2, wb["Men CPD"], 9, 8)

# Get month names with year
current_month_name = get_month_name_with_year(current_month_num, current_year)
month_1_before_name = get_month_name_with_year(month_1_before_num, month_1_before_year)
month_2_before_name = get_month_name_with_year(month_2_before_num, month_2_before_year)

# Adding month names to the specified cells directly
sheet = wb["Men CPD"]
sheet.cell(row=7, column=5, value=current_month_name)
sheet.cell(row=7, column=4, value=month_1_before_name)
sheet.cell(row=7, column=3, value=month_2_before_name)
sheet.cell(row=7, column=10, value=current_month_name)
sheet.cell(row=7, column=9, value=month_1_before_name)
sheet.cell(row=7, column=8, value=month_2_before_name)
sheet.cell(row=4, column=13, value=current_month_name)


# Men CPD Market Summary Table
# this month
load_data_into_template(function_df_asc, wb["Men CPD"], 11, 15)
load_data_into_template(format_df_asc, wb["Men CPD"], 15, 15)
# load_data_into_template(top_brands_df, wb["Men CPD"], 22, 15)
load_data_into_template(top_brands_df, wb["Men CPD"], 21, 15)

# Place the final_sentence in the 2nd row and 2nd column using the load_data_into_template function
load_data_into_template(final_sentence_df1, wb["Men CPD"], 2, 2)

# Create DataFrames to load into the template
sentence0_df = pd.DataFrame([[sentence0]])
sentence1_df = pd.DataFrame([[sentence1]])
growth_sentence_df = pd.DataFrame([[growth_sentence]])
sentence_df = pd.DataFrame([[sentence]])

# Use the load_data_into_template function to insert sentences into the correct cells
load_data_into_template(sentence0_df, sheet, 29, 2)  # Row 17, Column B (2)
load_data_into_template(sentence1_df, sheet, 30, 2)  # Row 18, Column B (2)
load_data_into_template(growth_sentence_df, sheet, 31, 2)  # Row 19, Column B (2)
load_data_into_template(sentence_df, sheet, 32, 2)  # Row 20, Column B (2)

# Save the workbook with changes
wb.save(template_path_commentary)

# %%
# Female CPD Summary Table
#this month
load_data_into_template(Market_Value3, wb["FEMALE + MUR (Mass Only)"], 8, 6)
load_data_into_template(Market_Unit3, wb["FEMALE + MUR (Mass Only)"], 8, 11)
load_data_into_template(Ms_Value3, wb["FEMALE + MUR (Mass Only)"], 9, 5)
load_data_into_template(Evo_Value3, wb["FEMALE + MUR (Mass Only)"], 9, 6)
load_data_into_template(Ms_Unit3, wb["FEMALE + MUR (Mass Only)"], 9, 10)
load_data_into_template(Evo_Unit3, wb["FEMALE + MUR (Mass Only)"], 9, 11)

#this month
load_data_into_template(Ms_Value4, wb["FEMALE + MUR (Mass Only)"], 9, 4)
load_data_into_template(Ms_Unit4, wb["FEMALE + MUR (Mass Only)"], 9, 9)

#this month
load_data_into_template(Ms_Value5, wb["FEMALE + MUR (Mass Only)"], 9, 3)
load_data_into_template(Ms_Unit5, wb["FEMALE + MUR (Mass Only)"], 9, 8)

# Get month names with year
current_month_name = get_month_name_with_year(current_month_num, current_year)
month_1_before_name = get_month_name_with_year(month_1_before_num, month_1_before_year)
month_2_before_name = get_month_name_with_year(month_2_before_num, month_2_before_year)

# Adding month names to the specified cells directly
sheet = wb["FEMALE + MUR (Mass Only)"]
sheet.cell(row=7, column=5, value=current_month_name)
sheet.cell(row=7, column=4, value=month_1_before_name)
sheet.cell(row=7, column=3, value=month_2_before_name)
sheet.cell(row=7, column=10, value=current_month_name)
sheet.cell(row=7, column=9, value=month_1_before_name)
sheet.cell(row=7, column=8, value=month_2_before_name)
sheet.cell(row=4, column=13, value=current_month_name)

# Market Summary Table
#this month
load_data_into_template(function_df_asc3, wb["FEMALE + MUR (Mass Only)"], 11, 15)
load_data_into_template(format_df_asc3, wb["FEMALE + MUR (Mass Only)"], 17, 15)
load_data_into_template(top_brands_df3, wb["FEMALE + MUR (Mass Only)"], 24, 15)


# Place the final_sentence in the 2nd row and 2nd column using the load_data_into_template function
load_data_into_template(final_sentence_df2, wb["FEMALE + MUR (Mass Only)"], 2, 2)

# Create DataFrames to load into the template
sentence_0_df = pd.DataFrame([[sentence_0]])
sentence_1_df = pd.DataFrame([[sentence_1]])
growth_sentence1_df = pd.DataFrame([[growth_sentence1]])
sentence_2_df = pd.DataFrame([[sentence_2]])

# Use the load_data_into_template function to insert sentences into the correct cells
load_data_into_template(sentence_0_df, sheet, 29, 2)  # Row 17, Column B (2)
load_data_into_template(sentence_1_df, sheet, 30, 2)  # Row 18, Column B (2)
load_data_into_template(growth_sentence1_df, sheet, 31, 2)  # Row 19, Column B (2)
load_data_into_template(sentence_2_df, sheet, 32, 2)  # Row 20, Column B (2)

# Save the workbook with changes
wb.save(template_path_commentary)

# %% [markdown]
# Save the final output
# 

# %%
# Get the current date
current_date = datetime.now()

# Calculate the previous month
previous_month = current_date - relativedelta(months=1)

# Get the name of the previous month
month_name = previous_month.strftime("%B")

# Create folder structure for the month inside the subfolder
month_subfolder_name = previous_month.strftime("%Y-%m")
subfolder_path = os.path.join(output_report_folder, month_subfolder_name)
os.makedirs(subfolder_path, exist_ok=True)

# Save the modified template to the correct month subfolder
base_template_name = os.path.splitext(os.path.basename(template_path_commentary))[0]
output_template_file_name = f"{base_template_name.replace('Template','')} ({month_name}).xlsx"
output_template_path = os.path.join(subfolder_path, output_template_file_name) 

# Save the modified workbook
wb.save(output_template_path)

# %%



