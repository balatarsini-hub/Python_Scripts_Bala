# %% [markdown]
# # SG LDB Commentary 

# %%
import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta

dir = os.getcwd()

# Get the current date
current_date = datetime.now()

# Calculate the previous month
previous_month = current_date - relativedelta(months=1)

# Create folder structure for the month inside the subfolder
month_subfolder_name = previous_month.strftime("%Y-%m")


#Specify folder for commentary
generated_report_folder = f'{dir}/Full Data'
commentary_template_folder = f'{dir}/Template/Commentary Template'
output_report_folder = f'{dir}/Generated Report/Commentary Report'
generated_report_folder

# %% [markdown]
# Load the file

# %%
# Choosing file
desired_generated_report_file = r"C:\Users\balatarsini_avinitya\Downloads\Topline Report - July\Full Data\SG CPD&LDB Female&Male Skincare Topline_Final 2025-12-29 Corn (1)_CLEAN_full.xlsx"

# Construct the full path to the specified generated data file
generated_report_path = os.path.join(generated_report_folder, desired_generated_report_file)

# Load all sheets of the generated data into a dictionary
generated_report = pd.read_excel(generated_report_path, sheet_name=None)

# Load the empty template from 'Commentary Template' with multiple sheets
template_path_commentary = os.path.join(commentary_template_folder, "SG Skincare LDB Commentary Template.xlsx")

# %% [markdown]
# # Important Sheets for Commentary

# %%
# Accessing a specific sheet from the dictionary
sheet_data = generated_report['12-Table-1']
sheet_data1= generated_report['13-Table-1']
sheet_data2 = generated_report['14-Table-1']
sheet_data3 = generated_report['15-Table-1']

# %% [markdown]
# LDB Summary Table

# %%
cpdsummary_df = sheet_data.iloc[17:84, 0:38]
cpdsummary_df1 = sheet_data2.iloc[17:84, 0:38]
cpdsummary_df2 = sheet_data1.iloc[:, 0:38]

# Retrieve and format the month and year
month_year = cpdsummary_df2.columns[35]
month_year_formatted = f"{month_year[:3]}'{month_year[-2:]} vs. {month_year[:3]}'{int(month_year[-2:])-1:02d}"

# Retrieve and format the market sales value
sales_value = cpdsummary_df2.iloc[1, 35]
if sales_value >= 1e6:
    sales_value_formatted = f"{sales_value / 1e6:.1f} mil"
else:
    sales_value_formatted = f"{sales_value / 1e3:.1f} thousand"

# Retrieve and format the evolution percentage
evo_percentage = cpdsummary_df2.iloc[1, 36]
evo_percentage_formatted = f"{evo_percentage:.1f}%"

# Construct the final sentence
final_sentence = f"{month_year_formatted}: Market sales at {sales_value_formatted} with {evo_percentage_formatted} evo."

# Create a DataFrame with the final_sentence
final_sentence_df = pd.DataFrame([[final_sentence]])

# list of brands to filter
brands_to_filter = ['TOTAL LDB', 'LA ROCHE POSAY', 'VICHY','Cerave']

# find the correct column name containing brand information
brand_column = cpdsummary_df.columns[0]
brand_column1 = cpdsummary_df1.columns[0]

# filter the DataFrame based on the specified brands
filtered_rows = cpdsummary_df[cpdsummary_df[brand_column].isin(brands_to_filter)]
filtered_rows1 = cpdsummary_df1[cpdsummary_df1[brand_column1].isin(brands_to_filter)]

# Value (this month)
Market_Value = sheet_data1.iloc[1:2, 36:37]/100
Ms_Value = filtered_rows.iloc[:,35:36]/100
Evo_Value = filtered_rows.iloc[:,36:37]/100

# Unit (this month)
Market_Unit =sheet_data3.iloc[1:2, 36:37]/100
Ms_Unit = filtered_rows1.iloc[:,35:36] /100
Evo_Unit = filtered_rows1.iloc[:,36:37]/100

# # Value (Last month)
Market_Value1 = sheet_data1.iloc[1:2, 34:35]/100
Ms_Value1 = filtered_rows.iloc[:,33:34]/100
Evo_Value1 = filtered_rows.iloc[:,34:35]/100

# Unit (Last month)
Market_Unit1 =sheet_data3.iloc[1:2, 34:35]/100
Ms_Unit1 = filtered_rows1.iloc[:,33:34] /100
Evo_Unit1 = filtered_rows1.iloc[:,34:35]/100

# Value (Last 2 month)
Market_Value2 = sheet_data1.iloc[1:2, 32:33]/100
Ms_Value2 = filtered_rows.iloc[:,31:32]/100
Evo_Value2 = filtered_rows.iloc[:,32:33]/100

# Unit (Last 2 month)
Market_Unit2 =sheet_data3.iloc[1:2, 32:33]/100
Ms_Unit2 = filtered_rows1.iloc[:,31:32] /100
Evo_Unit2 = filtered_rows1.iloc[:,32:33]/100

filtered_rows


# %% [markdown]
# For Date on the tables
# 

# %%
import re
from datetime import datetime, timedelta
import calendar

# Function to extract month and year
def extract_month_year(month_year):
    match = re.match(r"(\w{3})\s*(\d{2})", month_year)
    if match:
        month_str, year_str = match.groups()
        month = datetime.strptime(month_str, "%b").month
        year = int('20' + year_str)  # Adjust if year needs to be in 1900s or 2000s
        return month, year
    return None, None

# Function to get month name with year
def get_month_name_with_year(month, year):
    month_str = calendar.month_abbr[month]
    year_str = str(year)[2:]  # Get last 2 digits of the year
    return f"{month_str}'{year_str}"

# Function to get the month before a given month and year
def get_month_before(month, year, num_months):
    # Handle month subtraction with year transition
    for _ in range(num_months):
        if month == 1:
            month = 12
            year -= 1
        else:
            month -= 1
    return month, year

# Retrieve and clean column names
current_month = cpdsummary_df2.columns[35]
month_1_before = cpdsummary_df2.columns[34]

# Extract month and year
current_month_num, current_year = extract_month_year(current_month)
month_1_before_num, month_1_before_year = extract_month_year(month_1_before)

# Debug output
print(f"Current month column: {current_month}, Month number: {current_month_num}, Year: {current_year}")
print(f"Month 1 before column: {month_1_before}, Month number: {month_1_before_num}, Year: {month_1_before_year}")

# Check if extracted values are valid
if current_month_num is None or current_year is None:
    raise ValueError("Current month or year could not be extracted.")
if month_1_before_num is None or month_1_before_year is None:
    raise ValueError("Month 1 before could not be extracted.")

# Get the month before
month_2_before_num, month_2_before_year = get_month_before(current_month_num, current_year, 2)

date1 = get_month_name_with_year(current_month_num, current_year)
date2 = get_month_name_with_year(month_1_before_num, month_1_before_year)
date3 = get_month_name_with_year(month_2_before_num, month_2_before_year)


# %% [markdown]
# Market Summary Table

# %%
#Function
#concatenate df 
combine = pd.concat([sheet_data, sheet_data2], axis=1)

# functionsummary_df = combine.iloc[2:7,0:74]
functionsummary_df = combine.iloc[2:8,0:74]

column_names_function = ['Function','2022','2022 Evo', '2023', '2023 Evo', 'YTD-1','YTD-1 Evo','YTD' ,'YTD Evo', 'MAT-1', 'MAT-1 Evo','MAT','MAT Evo','Month1','Month1 Evo','Month2', 'Month2 Evo', 
                        'Month3','Month3 Evo','Month4','Month4 Evo', 'Month5','Month5 Evo', 'Month6','Month6 Evo','Month7','Month7 Evo','Month8','Month8 Evo','Month9','Month9 Evo','Month10','Month10 Evo',
                        'Month11','Month11 Evo','Month12','Month12 Evo','Function1','2022 Unit','2022 Unit Evo', '2023Unit','2023 Unit Evo','YTD-1 Unit','YTD-1 Unit Evo','YTD Unit' ,'YTD Unit Evo','MAT-1 Unit', 
                        'MAT-1 Unit Evo','MAT Unit','MAT Unit Evo','Month1 Unit','Month1 Unit Evo','Month2 Unit' ,'Month2 Unit Evo','Month3 Unit','Month3 Unit Evo','Month4 Unit','Month4 Unit Evo', 'Month5 Unit',
                        'Month5 Unit Evo', 'Month6 Unit','Month6 Unit Evo','Month7 Unit','Month7 Unit Evo','Month8 Unit','Month8 Unit Evo','Month9 Unit','Month9 Unit Evo','Month10 Unit','Month10 Unit Evo',
                        'Month11 Unit','Month11 Unit Evo','Month12 Unit','Month12 Unit Evo']
functionsummary_df.columns = column_names_function

replacements = {'WHITEN': 'Female Brightening'}

# %%
#Selecting data for this month
function_df_asc = functionsummary_df[~functionsummary_df['Function'].isin(['BASIC', 'HYDRATING'])][['Function', 'Month12 Evo', 'Month12 Unit Evo']].sort_values(by='Month12 Evo', ascending=False)
function_df_asc['Month12 Evo'] /= 100
function_df_asc['Month12 Unit Evo'] /= 100
function_df_asc['Function'] = function_df_asc['Function'].replace(replacements)

print(function_df_asc)

# %%
#Format
# formatsummary_df = combine.iloc[8:17,0:74]
formatsummary_df = combine.iloc[8:18,0:74]
formatsummary_df.columns = column_names_function

replacements = {'SERUM': 'Serum',
                'MAKE UP REMOVER': 'Make Up Remover',
                'SCRUB': 'Scrub',
                'OTHERS': 'Others'}

#select data this month
format_df_asc = formatsummary_df[~formatsummary_df['Function'].isin(['MOISTURIZER LOTION','MOISTURIZER WATER','TONER'])][['Function', 'Month12 Evo', 'Month12 Unit Evo']].sort_values(by = 'Month12 Evo', ascending = False)
format_df_asc['Month12 Evo'] /= 100
format_df_asc['Month12 Unit Evo'] /= 100
format_df_asc['Function'] = format_df_asc['Function'].replace(replacements)

# %%
formatsummary_df

# %%
#Brand
brandsummary_df = combine.iloc[20:84,0:74]
brandsummary_df.columns = column_names_function
# brand names
brand_names = ['LA ROCHE POSAY', 'VICHY', 'AVENE', 'Neutrogena', 'Eucerin', 'CETAPHIL', 'BIODERMA', 'BIONIKE', 'CUREL', 'EVANS', 'PLACENTOR VEGETAL', 
               'SEBAMED', 'URIAGE', 'Cerave', 'EGO', 'MUSTELA', 'PHYSIOGEL', 'QV', 'TOPICREM', 'OTHERS', 'EXCLUSIVE BRANDS', 'Nanowhite']

def identify_brand(brand):
    if brand in brand_names:
        return 'brand'
    else:
        return 'aux'

# Apply the function to the 'Brand' column
brandsummary_df['Brand_Type'] = brandsummary_df['Function'].apply(identify_brand)

replacements = {'LA ROCHE POSAY': 'La Roche Posay',
                'VICHY': 'Vichy',
                'AVENE': 'Avene',
                'Neutrogena': 'Neutrogena',
                'Eucerin': 'Eucerin',
                'CETAPHIL': 'Cetaphil',
                'BIODERMA': 'Bioderma',
                'BIONIKE': 'Bionike',
                'CUREL': 'Curel',
                'EVANS': 'Evans',
                'PLACENTOR VEGETAL': 'Placentor Vegetal',
                'SEBAMED': 'Sebamed',
                'URIAGE': 'Uriage',
                'EGO': 'Ego',
                'MUSTELA': 'Mustela',
                'PHYSIOGEL': 'Physiogel',
                'QV': 'QV',
                'TOPICREM': 'Topicrem',
                'OTHERS': 'Others',
                'EXCLUSIVE BRANDS': 'Exclusive Brands'}

# %%
# Select data for this month
top_brands_df = brandsummary_df[brandsummary_df['Brand_Type'] == 'brand'][['Function', 'Month12 Evo', 'Month12 Unit Evo']].sort_values(by = 'Month12 Evo', ascending = False).head(3)
top_brands_df['Month12 Evo'] /= 100
top_brands_df['Month12 Unit Evo'] /= 100
top_brands_df['Function'] = top_brands_df['Function'].replace(replacements)
top_brands_df['Function']

# %%
def ahead_or_behind(growth_value, value):
    c, m = growth_value, value

    if m < 0 and c < 0:
        # both negative
        if c < m:      # company more negative
            return "behind"    # (1)
        elif c > m:    # company less negative
            return "ahead"     # (2)
        else:
            return "inline"

    if m < 0 and c >= 0:
        return "ahead"          # (3)

    if m > 0 and c > 0:
        if c > m:
            return "ahead"      # (4)
        elif c < m:
            return "behind"     # (5)
        else:
            return "inline"

    if m > 0 and c <= 0:
        return "behind"         # (6)

    # market == 0 fallback
    if m == 0:
        if c > 0:  return "ahead"
        if c < 0:  return "behind"
        return "inline"

# %% [markdown]
# For YTD sentences at the bottom

# %%
# Assuming `date1` is already defined and contains the formatted month and year
sentence0 = f"YTD {date1}"
print(sentence0)

# Access the value from row 3, column I (8th column, as 'A' is 0)
value = sheet_data.iloc[1, 8]  # Use iloc to access row 3 (index 1) and column I (index 8)

# Determine if the market grew or declined
if value is not None:
    rounded_value = round(value, 1)  # Round the value to one decimal place
    if rounded_value > 0:
        sentence1 = f"1. Market grew +{rounded_value}% vs LY"
    else:
        sentence1 = f"1. Market declined {rounded_value}% vs LY"
else:
    sentence1 = "1. Value in cell 3I is not available."

print(sentence1)

# Access the values from the required cells
growth_value = sheet_data.iloc[18, 8]  # Value from cell 22I (index 21, column index 8)
market_comparison_value = sheet_data.iloc[1, 8]  # Value from cell 2I (index 1, column index 8)
market_share_current = sheet_data.iloc[18, 7]  # Value from cell 22H (index 21, column index 7)
market_share_ly = sheet_data.iloc[18, 5]  # Value from cell 22F (index 21, column index 5)

# Calculate rounded growth value and ratio for "ahead/behind" calculation
rounded_growth_value = round(growth_value, 1)
ratio_ahead_behind = round(growth_value / value, 1) if value != 0 else 0.0
ahead_behind = ahead_or_behind(growth_value, value)

# Calculate the market share difference
market_share_difference = round(market_share_current - market_share_ly, 1)

# Round all values to one decimal place for consistent formatting
market_share_current_rounded = round(market_share_current, 1)

# Generate the growth sentence
if rounded_growth_value > 0:
    growth_sentence = (
        f"2. Loreal LDB grew +{rounded_growth_value:.1f}%, "
        f"{abs(ratio_ahead_behind):.1f}x {ahead_behind} of market with "
        f"{market_share_current_rounded:.1f}%MS ({market_share_difference:+.1f} pp MS vs LY)"
    )
else:
    growth_sentence = (
        f"2. Loreal LDB declined {rounded_growth_value:.1f}%, "
        f"{abs(ratio_ahead_behind):.1f}x {ahead_behind} of market with "
        f"{market_share_current_rounded:.1f}%MS ({market_share_difference:+.1f} pp MS vs LY)"
    )

print(growth_sentence)

# List of brands to include in the final sentence
brands_of_interest = [
    "La Roche Posay",
    "Vichy",
    "Cerave",
    "Avene",
    "Neutrogena",
    "Eucerin",
    "Cetaphil",
    "Bioderma",
    "Bionike",
    "Curel",
    "Evans",
    "Placentor Vegetal",
    "Sebamed",
    "Uriage",
    "Ego",
    "Mustela",
    "Physiogel",
    "Qv",
    "Topicrem",
    "Others",
    "Exclusive Brands",
]

# Access the DataFrame for the specified rows and columns, filtering by the brands of interest
brands = (
    sheet_data.iloc[20:85, 0]        # adjust to 21:85 if you truly mean rows 22–85 (1-based)
    .dropna()
    .astype(str)
    .str.strip()
    .str.title()
    .tolist()
)  # Brands from row 22 to row 85 (index 21 to 84)

column_h_values = sheet_data.iloc[20:85, 7]  # Values from column H
column_f_values = sheet_data.iloc[20:85, 5]  # Values from column F

# Initialize variables to keep track of the best result
max_difference = float('-inf')
best_brand = None

# Loop through the brands to find the one with the highest difference
for i, brand in enumerate(brands):
    if brand in brands_of_interest:  # Check if the brand is in the list of brands of interest
        difference = column_h_values.iloc[i] - column_f_values.iloc[i]  # Calculate difference
        if difference > max_difference:
            max_difference = difference
            best_brand = brand

# Generate the sentence with the highest difference
if best_brand is not None:
    rounded_difference = round(max_difference, 1)  # Round the difference to one decimal place
    sentence = f"3. Share gain led by {best_brand} (+{rounded_difference:.1f} pp MS vs LY)"
else:
    sentence = "3. No share gain data available."

print(sentence)


# %% [markdown]
# Load Data into Template

# %%
# Load the template workbook
wb = load_workbook(template_path_commentary)

# Function to load selected data into the template sheet
def load_data_into_template(data, sheet, start_row, start_column):
    for i, row in enumerate(data.values, start=start_row):
        for j, value in enumerate(row, start=start_column):
            sheet.cell(row=i, column=j, value=value)

# LDB Summary Table
#this month
load_data_into_template(Market_Value, wb["Sheet1"], 8, 6)
load_data_into_template(Market_Unit, wb["Sheet1"], 8, 11)
load_data_into_template(Ms_Value, wb["Sheet1"], 9, 5)
load_data_into_template(Evo_Value, wb["Sheet1"], 9, 6)
load_data_into_template(Ms_Unit, wb["Sheet1"], 9, 10)
load_data_into_template(Evo_Unit, wb["Sheet1"], 9, 11)

#last month
load_data_into_template(Ms_Value1, wb["Sheet1"], 9, 4)
load_data_into_template(Ms_Unit1, wb["Sheet1"], 9, 9)

#last 2 month
load_data_into_template(Ms_Value2, wb["Sheet1"], 9, 3)
load_data_into_template(Ms_Unit2, wb["Sheet1"], 9, 8)

# Get month names with year
current_month_name = get_month_name_with_year(current_month_num, current_year)
month_1_before_name = get_month_name_with_year(month_1_before_num, month_1_before_year)
month_2_before_name = get_month_name_with_year(month_2_before_num, month_2_before_year)

# Adding month names to the specified cells directly
sheet = wb["Sheet1"]
sheet.cell(row=7, column=5, value=current_month_name)
sheet.cell(row=7, column=4, value=month_1_before_name)
sheet.cell(row=7, column=3, value=month_2_before_name)
sheet.cell(row=7, column=10, value=current_month_name)
sheet.cell(row=7, column=9, value=month_1_before_name)
sheet.cell(row=7, column=8, value=month_2_before_name)
sheet.cell(row=4, column=13, value=current_month_name)

# Market Summary Table

#this month
load_data_into_template(function_df_asc, wb["Sheet1"], 11, 15)
load_data_into_template(format_df_asc, wb["Sheet1"], 16, 15)
load_data_into_template(top_brands_df, wb["Sheet1"], 23, 15)

# Place the final_sentence in the 2nd row and 2nd column using the load_data_into_template function
load_data_into_template(final_sentence_df, wb["Sheet1"], 2, 2)

# Create DataFrames to load into the template
sentence0_df = pd.DataFrame([[sentence0]])
sentence1_df = pd.DataFrame([[sentence1]])
growth_sentence_df = pd.DataFrame([[growth_sentence]])
sentence_df = pd.DataFrame([[sentence]])

# Use the load_data_into_template function to insert sentences into the correct cells
load_data_into_template(sentence0_df, sheet, 35, 2)  # Row 17, Column B (2)
load_data_into_template(sentence1_df, sheet, 36, 2)  # Row 18, Column B (2)
load_data_into_template(growth_sentence_df, sheet, 37, 2)  # Row 19, Column B (2)
load_data_into_template(sentence_df, sheet, 38, 2)  # Row 20, Column B (2)

# Save the workbook with changes
wb.save(template_path_commentary)

# %% [markdown]
# Save the final output

# %%
# Get the current date
current_date = datetime.now()

# Calculate the previous month
previous_month = current_date - relativedelta(months=1)

# Get the name of the previous month
month_name = previous_month.strftime("%B")

# Create folder structure for the month inside the subfolder
month_subfolder_name = previous_month.strftime("%Y-%m")
subfolder_path = os.path.join(output_report_folder, month_subfolder_name)
os.makedirs(subfolder_path, exist_ok=True)

# Save the modified template to the correct month subfolder
base_template_name = os.path.splitext(os.path.basename(template_path_commentary))[0]
output_template_file_name = f"{base_template_name.replace('Template','')} ({month_name}).xlsx"
output_template_path = os.path.join(subfolder_path, output_template_file_name) 

# Save the modified workbook
wb.save(output_template_path)


