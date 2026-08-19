from __future__ import annotations

import argparse
import calendar
import csv
from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill


DEFAULT_CPD_DIR = Path(__file__).resolve().parent / "CPD"
DEFAULT_REPORT_DIR = Path(__file__).resolve().parent / "Hand and Body Pulse reports"
DEFAULT_OO_FILE = r"C:\Users\balatarsini_avinitya\Downloads\Final folders\Offline Bodycare ingestion V 23.06.26\Offline Bodycare ingestion\CPD\MYSG ONE CPD CMI YTD Jun'26.xlsx"
DEFAULT_SPLIT_FILE = "Bodycare split method and estimation.xlsx"
DEFAULT_MY_REPORT = r"C:\Users\balatarsini_avinitya\Downloads\Final folders\Offline Bodycare ingestion V 23.06.26\Offline Bodycare ingestion\Hand and Body Pulse reports\MY L'OREAL_Hand & Body Mositurizer Report - 2026-04-20 (1).xlsx"

OO_SHEET = "O+O Data"
REPORT_SHEET = "1-Total Category & LOreal"
MASS_REPORT_SHEET = "3-Total Mass & Top 10 brands"
REPORT_HEADER_ROW = 10
REPORT_PRODUCT_COL = 1
SOURCE_CATEGORIES = ("Female Skincare", "Male Skincare")
SOURCE_CATEGORY_LABEL = "Female + Male Skincare"
OUTPUT_CATEGORY = "Bodycare"
SOURCE_START_YEAR = 2024

BRAND_SOURCE_MAP = {
    "BRAND_02": "GARNIER",
}

OUTPUT_BRAND_MAP = {
    "BRAND_02": "Garnier",
    "Market": "Market",
}

REPORT_PRODUCTS = {
    "BRAND_02": (REPORT_SHEET, "GARNIER"),
    "Market": (MASS_REPORT_SHEET, "TOTAL MASS"),
}

FINAL_OUTPUT_ORDER = ["BRAND_02", "Market"]
FINAL_HEADERS = [
    "MY/SG",
    "Platform",
    "Year",
    "Month",
    "Brand_x",
    "Category_x",
    "Value",
]


def parse_target_month(value: str) -> tuple[int, int]:
    try:
        year_text, month_text = value.split("-", 1)
        year = int(year_text)
        month = int(month_text)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Use YYYY-MM, for example 2026-04") from exc
    if month < 1 or month > 12:
        raise argparse.ArgumentTypeError("Month must be between 01 and 12")
    return year, month


def month_label(year: int, month: int) -> str:
    return f"{calendar.month_abbr[month]}'{str(year)[-2:]}"


def quote_sheet(name: str) -> str:
    return name.replace("'", "''")


def external_ref(workbook_name: str, sheet_name: str, cell_ref: str) -> str:
    workbook_name = workbook_name.replace("'", "''")
    return f"'[{workbook_name}]{quote_sheet(sheet_name)}'!${cell_ref}"


def external_formula(workbook_name: str, sheet_name: str, cell_ref: str) -> str:
    return f"={external_ref(workbook_name, sheet_name, cell_ref)}"


def excel_col(number: int) -> str:
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result


def period_key(year: int, month: int) -> int:
    return year * 12 + month


def iter_months(start_year: int, target_year: int, target_month: int):
    for year in range(start_year, target_year + 1):
        month_end = target_month if year == target_year else 12
        for month in range(1, month_end + 1):
            yield year, month


def previous_month(year: int, month: int) -> tuple[int, int]:
    if month == 1:
        return year - 1, 12
    return year, month - 1


def quarter_start(month: int) -> int:
    return ((month - 1) // 3) * 3 + 1


def resolve_path(base_dir: Path, value: str | Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else base_dir / path


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.number_format:
            target.number_format = source.number_format


def get_groups(ws) -> list[tuple[int, int, str, str]]:
    groups = []
    start = 2
    for row in range(3, ws.max_row + 2):
        same_group = False
        if row <= ws.max_row:
            same_group = (
                ws.cell(row, 5).value == ws.cell(row - 1, 5).value
                and ws.cell(row, 6).value == ws.cell(row - 1, 6).value
            )
        if not same_group:
            brand = ws.cell(start, 5).value
            category = ws.cell(start, 6).value
            if brand and category:
                groups.append((start, row - 1, str(brand), str(category)))
            start = row
    return groups


def build_row_lookup(ws) -> dict[tuple[str, str, int, int], int]:
    lookup = {}
    for row in range(2, ws.max_row + 1):
        brand = ws.cell(row, 5).value
        category = ws.cell(row, 6).value
        year = ws.cell(row, 3).value
        month = ws.cell(row, 4).value
        if brand and category and isinstance(year, int) and isinstance(month, int):
            lookup[(str(brand), str(category), year, month)] = row
    return lookup


def find_report_rows_and_quarters(report_path: Path) -> tuple[dict[str, tuple[str, int]], dict[tuple[int, int], str]]:
    wb = load_workbook(report_path, data_only=True, read_only=True)
    row_by_brand = {}

    for brand, (sheet_name, expected_product) in REPORT_PRODUCTS.items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"{report_path.name}: missing report sheet {sheet_name!r}")
        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            product = ws.cell(row, REPORT_PRODUCT_COL).value
            if str(product or "").strip().upper() == expected_product.upper():
                row_by_brand[brand] = (sheet_name, row)
                break

    missing = sorted(set(REPORT_PRODUCTS) - set(row_by_brand))
    if missing:
        raise ValueError(f"{report_path.name}: missing report rows for {', '.join(missing)}")

    col_by_quarter = {}
    ws = wb[REPORT_SHEET]
    for col in range(2, ws.max_column + 1):
        if str(ws.cell(REPORT_HEADER_ROW - 1, col).value or "").strip() != "Sales Value":
            continue
        header = ws.cell(REPORT_HEADER_ROW, col).value
        if not header:
            continue
        text = str(header)
        if not text.startswith("Q"):
            continue
        quarter = int(text[1])
        year = 2000 + int(text.split()[1])
        col_by_quarter[(year, quarter)] = excel_col(col)
    return row_by_brand, col_by_quarter


def read_report_value(
    report_wb,
    report_rows: dict[str, tuple[str, int]],
    report_quarters: dict[tuple[int, int], str],
    brand: str,
    year: int,
    quarter: int,
) -> float:
    sheet_name, row = report_rows[brand]
    col_letter = report_quarters[(year, quarter)]
    value = report_wb[sheet_name][f"{col_letter}{row}"].value
    if value is None:
        raise ValueError(f"Missing pulse report value for {brand} {year} Q{quarter}")
    return float(value)


def same_country_currency(country_currency: str | None, country: str) -> bool:
    return str(country_currency or "").startswith(f"{country} ")


def load_oo_monthly_values(
    oo_path: Path,
    source_categories: tuple[str, ...],
    start_year: int,
) -> tuple[dict[tuple[str, str, int, int], float], dict[tuple[str, str, str, int, int], float], tuple[int, int]]:
    wb = load_workbook(oo_path, data_only=True, read_only=True)
    if OO_SHEET not in wb.sheetnames:
        raise ValueError(f"{oo_path.name}: missing sheet {OO_SHEET!r}")
    ws = wb[OO_SHEET]

    monthly_values = defaultdict(float)
    channel_values = defaultdict(float)
    periods = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        country_currency = row[0]
        loreal_or_market = row[1]
        channel = row[2]
        offline_online = str(row[3] or "").strip()
        raw_brand = str(row[4] or "").strip().upper()
        value = row[6] or 0
        country = row[8]
        year = row[10]
        month = row[11]
        brand_x = str(row[12] or "").strip()
        category_x = str(row[13] or "").strip()

        if country not in {"MY", "SG"}:
            continue
        if not same_country_currency(country_currency, country):
            continue
        if not isinstance(year, int) or not isinstance(month, int) or year < start_year:
            continue
        if offline_online != "Offline":
            continue
        if category_x not in source_categories:
            continue

        output_brand = None
        if loreal_or_market == "Market" and raw_brand == "EXC. MASS MEDIC":
            output_brand = "Market"
        elif loreal_or_market == "L'Oreal":
            for workbook_brand, source_brand in BRAND_SOURCE_MAP.items():
                if brand_x.upper() == source_brand:
                    output_brand = workbook_brand
                    break

        if output_brand is None:
            continue

        numeric_value = float(value)
        monthly_values[(country, output_brand, year, month)] += numeric_value
        channel_values[(country, output_brand, str(channel or ""), year, month)] += numeric_value
        periods.append((year, month))

    if not periods:
        raise ValueError(f"{oo_path.name}: no matching offline O+O rows found for {source_categories!r}")
    return dict(monthly_values), dict(channel_values), max(periods)


def build_split_method_workbook(
    oo_path: Path,
    output_path: Path,
    source_categories: tuple[str, ...],
    start_year: int,
) -> tuple[int, int, dict[tuple[str, str, int, int], float]]:
    monthly_values, channel_values, max_period = load_oo_monthly_values(oo_path, source_categories, start_year)
    periods = sorted({(year, month) for _, _, _, year, month in channel_values})

    wb = Workbook()
    market_ws = wb.active
    market_ws.title = "Market_01"
    brand_ws = wb.create_sheet("Brand_02")

    for ws in [market_ws, brand_ws]:
        ws.cell(2, 1).value = "Brand"
        ws.cell(2, 2).value = "Category"
        ws.cell(2, 3).value = "Channel"
        for offset, (year, month) in enumerate(periods, start=4):
            ws.cell(1, offset).value = year
            ws.cell(2, offset).value = month

    rows_by_sheet = {
        "Market_01": ("Market", "EXC. MASS MEDIC"),
        "Brand_02": ("BRAND_02", "Brand_02"),
    }
    for ws in [market_ws, brand_ws]:
        workbook_brand, display_brand = rows_by_sheet[ws.title]
        output_row = 3
        channels = sorted(
            {
                channel
                for country, brand, channel, _, _ in channel_values
                if brand == workbook_brand and country == "MY"
            }
        )
        for channel in channels:
            ws.cell(output_row, 1).value = display_brand
            ws.cell(output_row, 2).value = SOURCE_CATEGORY_LABEL
            ws.cell(output_row, 3).value = channel
            for offset, (year, month) in enumerate(periods, start=4):
                ws.cell(output_row, offset).value = channel_values.get(("MY", workbook_brand, channel, year, month), 0)
            output_row += 1

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for ws in [market_ws, brand_ws]:
        for row in [1, 2]:
            for cell in ws[row]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
        for column_cells in ws.columns:
            ws.column_dimensions[column_cells[0].column_letter].width = 15

    wb.save(output_path)
    return max_period[0], max_period[1], monthly_values


def split_sum_formula(split_file_name: str, sheet_name: str, year: int, month: int, split_path: Path) -> str:
    wb = load_workbook(split_path, data_only=False, read_only=True)
    ws = wb[sheet_name]
    period_col = None
    for col in range(4, ws.max_column + 1):
        if ws.cell(1, col).value == year and ws.cell(2, col).value == month:
            period_col = ws.cell(1, col).column_letter
            break
    if period_col is None:
        raise ValueError(f"{split_path.name}: missing period {year}-{month:02d} in {sheet_name}")
    first_row = 3
    last_row = ws.max_row
    workbook_name = split_file_name.replace("'", "''")
    prefix = f"'[{workbook_name}]{quote_sheet(sheet_name)}'!"
    return f"=SUM({prefix}${period_col}${first_row}:${period_col}${last_row})"


def ensure_months_through(ws, target_year: int, target_month: int, output_category: str) -> None:
    while True:
        groups = [group for group in get_groups(ws) if group[3] == output_category]
        row_lookup = build_row_lookup(ws)
        missing = []
        complete = True
        for _, end, brand, category in groups:
            if (brand, category, target_year, target_month) in row_lookup:
                continue
            last_year = int(ws.cell(end, 3).value)
            last_month = int(ws.cell(end, 4).value)
            next_year, next_month = (last_year + 1, 1) if last_month == 12 else (last_year, last_month + 1)
            missing.append((end, next_year, next_month))
            complete = False

        if complete:
            return

        for previous_row, next_year, next_month in sorted(missing, reverse=True):
            ws.insert_rows(previous_row + 1, 1)
            copy_row_style(ws, previous_row, previous_row + 1)
            for col in [1, 2, 5, 6, 13]:
                ws.cell(previous_row + 1, col).value = ws.cell(previous_row, col).value
            ws.cell(previous_row + 1, 3).value = next_year
            ws.cell(previous_row + 1, 4).value = next_month


def sum_range(rows: list[int], col: str = "G") -> str:
    if not rows:
        return "0"
    return f"SUM({col}{min(rows)}:{col}{max(rows)})"


def set_if_writable(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def quarter_rows(row_lookup, brand: str, category: str, year: int, q_start: int) -> list[int]:
    return [
        row_lookup[(brand, category, year, month)]
        for month in range(q_start, q_start + 3)
        if (brand, category, year, month) in row_lookup
    ]


def update_bodycare_formulas(
    ws,
    split_path: Path,
    split_file_name: str,
    report_file_name: str,
    report_rows: dict[str, tuple[str, int]],
    report_quarters: dict[tuple[int, int], str],
    target_year: int,
    target_month: int,
    split_max_year: int,
    split_max_month: int,
    output_category: str,
) -> None:
    ensure_months_through(ws, target_year, target_month, output_category)
    row_lookup = build_row_lookup(ws)
    split_max_key = period_key(split_max_year, split_max_month)

    for start, end, brand, category in get_groups(ws):
        if category != output_category:
            continue
        sheet_name = "Market_01" if brand == "Market" else "Brand_02"
        for row in range(start, end + 1):
            year = int(ws.cell(row, 3).value)
            month = int(ws.cell(row, 4).value)
            if period_key(year, month) > period_key(target_year, target_month):
                continue

            q_start = quarter_start(month)
            quarter = (month - 1) // 3 + 1
            has_pulse = (year, quarter) in report_quarters
            report_sheet, report_row = report_rows[brand]

            if period_key(year, month) <= split_max_key:
                ws.cell(row, 7).value = split_sum_formula(split_file_name, sheet_name, year, month, split_path)
            else:
                prev_year, prev_month = previous_month(year, month)
                previous_row = row_lookup[(brand, category, prev_year, prev_month)]
                prior_year_row = row_lookup[(brand, category, year - 1, month)]
                ws.cell(row, 7).value = f"=G{previous_row}*(1+H{prior_year_row})"

            ws.cell(row, 8).value = None if row == start else f"=G{row}/G{row - 1}-1"

            rows = quarter_rows(row_lookup, brand, category, year, q_start)
            ws.cell(row, 9).value = f"=G{row}/{sum_range(rows)}"
            should_hold_actual = period_key(year, month) <= split_max_key and has_pulse and month == q_start
            if should_hold_actual:
                col = report_quarters[(year, quarter)]
                set_if_writable(ws, row, 10, external_formula(report_file_name, report_sheet, f"{col}{report_row}"))
            else:
                set_if_writable(ws, row, 10, None)

            k_row_current = row_lookup.get((brand, category, year, q_start))
            previous_period = previous_month(year, month)
            previous_row = row_lookup.get((brand, category, previous_period[0], previous_period[1]))
            if has_pulse and k_row_current:
                set_if_writable(ws, row, 11, f"=$J${k_row_current}*I{row}")
            elif previous_row:
                set_if_writable(ws, row, 11, f"=K{previous_row}/(1+H{row})")
            else:
                set_if_writable(ws, row, 11, f"=G{row}")
            set_if_writable(ws, row, 13, report_sheet)

            ws.cell(row, 8).number_format = "0.00%"
            ws.cell(row, 9).number_format = "0.00%"
            for col in [7, 10, 11]:
                ws.cell(row, col).number_format = "#,##0.00"


def calculate_bodycare_values(
    ws,
    monthly_values: dict[tuple[str, str, int, int], float],
    report_path: Path,
    report_rows: dict[str, tuple[str, int]],
    report_quarters: dict[tuple[int, int], str],
    target_year: int,
    target_month: int,
    split_max_year: int,
    split_max_month: int,
    output_category: str,
) -> dict[int, float]:
    report_wb = load_workbook(report_path, data_only=True, read_only=True)
    row_lookup = build_row_lookup(ws)
    split_max_key = period_key(split_max_year, split_max_month)
    h_values = {}
    evo = {}
    shares = {}
    actuals = {}
    final_values = {}

    for start, end, brand, category in get_groups(ws):
        if category != output_category:
            continue
        for row in range(start, end + 1):
            year = int(ws.cell(row, 3).value)
            month = int(ws.cell(row, 4).value)
            if period_key(year, month) > period_key(target_year, target_month):
                continue
            if period_key(year, month) <= split_max_key:
                h_values[row] = monthly_values.get(("MY", brand, year, month), 0.0)
            else:
                prev_year, prev_month = previous_month(year, month)
                previous_row = row_lookup[(brand, category, prev_year, prev_month)]
                prior_year_row = row_lookup[(brand, category, year - 1, month)]
                h_values[row] = h_values[previous_row] * (1 + evo[prior_year_row])
            evo[row] = None if row == start else h_values[row] / h_values[row - 1] - 1

        for row in range(start, end + 1):
            year = int(ws.cell(row, 3).value)
            month = int(ws.cell(row, 4).value)
            if period_key(year, month) > period_key(target_year, target_month):
                continue

            q_start = quarter_start(month)
            quarter = (month - 1) // 3 + 1
            q_rows = quarter_rows(row_lookup, brand, category, year, q_start)
            denominator = sum(h_values[q_row] for q_row in q_rows if q_row in h_values)
            shares[row] = h_values[row] / denominator if denominator else 0

            has_pulse = (year, quarter) in report_quarters
            should_hold_actual = period_key(year, month) <= split_max_key and has_pulse and month == q_start
            actuals[row] = (
                read_report_value(report_wb, report_rows, report_quarters, brand, year, quarter)
                if should_hold_actual
                else None
            )

            k_row_current = row_lookup.get((brand, category, year, q_start))
            previous_period = previous_month(year, month)
            previous_row = row_lookup.get((brand, category, previous_period[0], previous_period[1]))
            if has_pulse and k_row_current:
                final_values[row] = actuals[k_row_current] * shares[row]
            elif previous_row:
                final_values[row] = final_values[previous_row] / (1 + evo[row])
            else:
                final_values[row] = h_values[row]
    return final_values


def calculate_bodycare_values_direct(
    monthly_values: dict[tuple[str, str, int, int], float],
    report_path: Path,
    report_rows: dict[str, tuple[str, int]],
    report_quarters: dict[tuple[int, int], str],
    start_year: int,
    target_year: int,
    target_month: int,
    split_max_year: int,
    split_max_month: int,
) -> dict[tuple[str, int, int], float]:
    report_wb = load_workbook(report_path, data_only=True, read_only=True)
    split_max_key = period_key(split_max_year, split_max_month)
    final_values = {}

    for brand in FINAL_OUTPUT_ORDER:
        h_values = {}
        evo = {}
        shares = {}
        actuals = {}

        for year, month in iter_months(start_year, target_year, target_month):
            key = (year, month)
            if period_key(year, month) <= split_max_key:
                h_values[key] = monthly_values.get(("MY", brand, year, month), 0.0)
            else:
                prev_key = previous_month(year, month)
                prior_year_key = (year - 1, month)
                if prev_key not in h_values or prior_year_key not in evo:
                    raise ValueError(
                        f"Cannot estimate {brand} {year}-{month:02d}; missing prior month/year values"
                    )
                h_values[key] = h_values[prev_key] * (1 + evo[prior_year_key])

            prev_key = previous_month(year, month)
            evo[key] = None if prev_key not in h_values else h_values[key] / h_values[prev_key] - 1

        for year, month in iter_months(start_year, target_year, target_month):
            key = (year, month)
            q_start = quarter_start(month)
            quarter = (month - 1) // 3 + 1
            q_keys = [(year, q_month) for q_month in range(q_start, q_start + 3)]
            denominator = sum(h_values[q_key] for q_key in q_keys if q_key in h_values)
            shares[key] = h_values[key] / denominator if denominator else 0

            has_pulse = (year, quarter) in report_quarters
            should_hold_actual = period_key(year, month) <= split_max_key and has_pulse and month == q_start
            actuals[key] = (
                read_report_value(report_wb, report_rows, report_quarters, brand, year, quarter)
                if should_hold_actual
                else None
            )

            quarter_key = (year, q_start)
            prev_key = previous_month(year, month)
            if has_pulse and quarter_key in h_values:
                if actuals.get(quarter_key) is None:
                    raise ValueError(f"Missing pulse actual for {brand} {year} Q{quarter}")
                final_values[(brand, year, month)] = actuals[quarter_key] * shares[key]
            elif prev_key in h_values:
                final_values[(brand, year, month)] = final_values[(brand, *prev_key)] / (1 + evo[key])
            else:
                final_values[(brand, year, month)] = h_values[key]

    return final_values


def export_picked_values(
    monthly_values: dict[tuple[str, str, int, int], float],
    output_path: Path,
) -> int:
    rows = []
    for (country, brand, year, month), value in sorted(monthly_values.items()):
        rows.append(
            {
                "Country": country,
                "Output Brand": brand,
                "Source Brand": "EXC. MASS MEDIC" if brand == "Market" else BRAND_SOURCE_MAP.get(brand, brand),
                "Source Category": SOURCE_CATEGORY_LABEL,
                "Offline/Online": "Offline",
                "Mass/Non-Mass": "Mass" if brand == "Market" else "Mass",
                "Year": year,
                "Month": month,
                "Value": value,
            }
        )
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


def export_final_output(
    output_path: Path,
    values_by_row: dict[int, float],
    ws,
    start_year: int,
    target_year: int,
    target_month: int,
    output_category: str,
) -> int:
    wb = Workbook()
    out_ws = wb.active
    out_ws.title = "MY"
    for col, value in enumerate(FINAL_HEADERS, start=1):
        out_ws.cell(1, col).value = value
        out_ws.cell(1, col).font = Font(bold=True)
        out_ws.column_dimensions[out_ws.cell(1, col).column_letter].width = 16

    row_lookup = build_row_lookup(ws)
    output_row = 2
    for year in range(start_year, target_year + 1):
        month_end = target_month if year == target_year else 12
        for month in range(1, month_end + 1):
            for brand in FINAL_OUTPUT_ORDER:
                source_row = row_lookup.get((brand, output_category, year, month))
                if source_row is None:
                    continue
                out_ws.cell(output_row, 1).value = "MY"
                out_ws.cell(output_row, 2).value = "Offline est"
                out_ws.cell(output_row, 3).value = year
                out_ws.cell(output_row, 4).value = month
                out_ws.cell(output_row, 5).value = OUTPUT_BRAND_MAP[brand].upper()
                out_ws.cell(output_row, 6).value = output_category
                out_ws.cell(output_row, 7).value = values_by_row[source_row]
                out_ws.cell(output_row, 7).number_format = "#,##0.00"
                output_row += 1

    wb.save(output_path)
    return output_row - 2


def export_final_output_direct(
    output_path: Path,
    values: dict[tuple[str, int, int], float],
    start_year: int,
    target_year: int,
    target_month: int,
    output_category: str,
) -> int:
    wb = Workbook()
    out_ws = wb.active
    out_ws.title = "MY"
    for col, value in enumerate(FINAL_HEADERS, start=1):
        out_ws.cell(1, col).value = value
        out_ws.cell(1, col).font = Font(bold=True)
        out_ws.column_dimensions[out_ws.cell(1, col).column_letter].width = 16

    output_row = 2
    for year, month in iter_months(start_year, target_year, target_month):
        for brand in FINAL_OUTPUT_ORDER:
            value = values.get((brand, year, month))
            if value is None:
                continue
            out_ws.cell(output_row, 1).value = "MY"
            out_ws.cell(output_row, 2).value = "Offline est"
            out_ws.cell(output_row, 3).value = year
            out_ws.cell(output_row, 4).value = month
            out_ws.cell(output_row, 5).value = OUTPUT_BRAND_MAP[brand].upper()
            out_ws.cell(output_row, 6).value = output_category
            out_ws.cell(output_row, 7).value = value
            out_ws.cell(output_row, 7).number_format = "#,##0.00"
            output_row += 1

    wb.save(output_path)
    return output_row - 2


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Create CPD bodycare monthly offline output from the O+O file."
    )
    parser.add_argument("target_month", nargs="?", type=parse_target_month, help="Target month, e.g. 2026-04")
    parser.add_argument("--cpd-dir", type=Path, default=DEFAULT_CPD_DIR)
    parser.add_argument("--oo-file", default=DEFAULT_OO_FILE)
    parser.add_argument("--split-file", default=DEFAULT_SPLIT_FILE)
    parser.add_argument("--input", default=None, help="Optional estimation workbook to update")
    parser.add_argument("--output", default=None, help="Optional updated estimation workbook name")
    parser.add_argument("--final-start-year", type=int, default=2024)
    parser.add_argument("--final-output", default=None, help="Defaults to CPD Data Output <Mon>'<YY>.xlsx")
    parser.add_argument("--export-picked-values", default=None)
    parser.add_argument("--my-report", default=DEFAULT_MY_REPORT)
    args = parser.parse_args()

    if args.target_month:
        target_year, target_month = args.target_month
    else:
        target_year, target_month = parse_target_month(input("Enter target month (YYYY-MM): ").strip())

    final_output_name = args.final_output or f"CPD Data Output {month_label(target_year, target_month)}.xlsx"

    cpd_dir = args.cpd_dir
    oo_path = cpd_dir / args.oo_file
    split_path = cpd_dir / args.split_file
    final_output_path = cpd_dir / final_output_name
    my_report_path = resolve_path(cpd_dir, args.my_report)

    for required_path in [oo_path, my_report_path]:
        if not required_path.exists():
            raise FileNotFoundError(f"Required workbook not found: {required_path}")
    if args.output and not args.input:
        raise ValueError("--output requires --input because the estimation workbook is now optional")

    split_max_year, split_max_month, monthly_values = build_split_method_workbook(
        oo_path,
        split_path,
        SOURCE_CATEGORIES,
        SOURCE_START_YEAR,
    )

    if args.export_picked_values:
        row_count = export_picked_values(monthly_values, cpd_dir / args.export_picked_values)
        print(f"Exported {row_count} O+O picked values: {cpd_dir / args.export_picked_values}")

    report_rows, report_quarters = find_report_rows_and_quarters(my_report_path)

    values = calculate_bodycare_values_direct(
        monthly_values,
        my_report_path,
        report_rows,
        report_quarters,
        SOURCE_START_YEAR,
        target_year,
        target_month,
        split_max_year,
        split_max_month,
    )
    final_rows = export_final_output_direct(
        final_output_path,
        values,
        args.final_start_year,
        target_year,
        target_month,
        OUTPUT_CATEGORY,
    )

    if args.input:
        input_path = cpd_dir / args.input
        output_path = cpd_dir / (args.output or args.input)
        if not input_path.exists():
            raise FileNotFoundError(f"Input workbook not found: {input_path}")

        wb = load_workbook(input_path)
        if "MY CPD" not in wb.sheetnames:
            raise ValueError("Missing sheet: MY CPD")
        update_bodycare_formulas(
            wb["MY CPD"],
            split_path,
            args.split_file,
            my_report_path.name,
            report_rows,
            report_quarters,
            target_year,
            target_month,
            split_max_year,
            split_max_month,
            OUTPUT_CATEGORY,
        )
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.save(output_path)
        print(f"Updated optional estimation workbook: {output_path}")

    print(f"Updated split method through {split_max_year}-{split_max_month:02d}: {split_path}")
    print(f"Updated final output with {final_rows} rows: {final_output_path}")


if __name__ == "__main__":
    main()
