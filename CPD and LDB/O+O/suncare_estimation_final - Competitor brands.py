from __future__ import annotations
import argparse
import calendar
import csv
import pandas as pd
from copy import copy
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font, PatternFill

DEFAULT_LDB_DIR = Path(__file__).resolve().parent / "LDB"
DEFAULT_REPORT_DIR = Path(__file__).resolve().parent / "Suncare reports"
DEFAULT_SPLIT_FILE = "Bodycare split method and estimation.xlsx"
DEFAULT_OO_FILE = "MYSG ONE LDB CMI YTD Apr'26.xlsx"
DEFAULT_OUTPUT_TEMPLATE = "Data Output format.xlsx"
DEFAULT_MY_REPORT = DEFAULT_REPORT_DIR / "NIQP MY SUNCARE 20240401_20260331 (1).xlsx"
DEFAULT_SG_REPORT = DEFAULT_REPORT_DIR / "NIQP SG SUNCARE 20240401_20260331 (1).xlsx"
MAPPING_PATH = r"C:\Users\balatarsini_avinitya\Downloads\Final folders\Offline Suncare ingestion v 23.06.26\Offline Suncare ingestion\LDB\Mapping.xlsx"

# ---------------------------------------------------------------------------
# Pulse report sheet / row config
# ---------------------------------------------------------------------------
MY_REPORT_SHEET = "1-Malaysia Key Account"
SG_REPORT_SHEET = "1-Modern Trade"
MY_MARKET_ROW = 52
SG_MARKET_ROW = 53
REPORT_HEADER_ROW = 11
REPORT_PRODUCT_COL = 1

BRAND_REPORT_ROWS: dict[str, dict[str, tuple[str, int]]] = {
    "MY": {
        "BRAND_07": (MY_REPORT_SHEET, 16),
        "Market":   (MY_REPORT_SHEET, MY_MARKET_ROW),
    },
    "SG": {
        "BRAND_06": (SG_REPORT_SHEET, 14),
        "BRAND_07": (SG_REPORT_SHEET, 16),
        "BRAND_08": (SG_REPORT_SHEET, 18),
        "Market":   (SG_REPORT_SHEET, SG_MARKET_ROW),
    },
}

# ---------------------------------------------------------------------------
# Split workbook row config
# Brand sheet: row 3 = MY BRAND_07, row 4 = SG BRAND_06,
#              row 5 = SG BRAND_07,  row 6 = SG BRAND_08
# ---------------------------------------------------------------------------
SPLIT_SOURCE_ROWS: dict[tuple[str, str, str | None], tuple[str, int]] = {
    ("MY", "BRAND_07", None):              ("Brand",  3),
    ("SG", "BRAND_06", None):              ("Brand",  4),
    ("SG", "BRAND_07", None):              ("Brand",  5),
    ("SG", "BRAND_08", None):              ("Brand",  6),
    ("MY", "Market", "Mass Medical"):      ("Market", 3),
    ("MY", "Market", "Non-Mass Medical"):  ("Market", 4),
    ("SG", "Market", "Mass Medical"):      ("Market", 5),
    ("SG", "Market", "Non-Mass Medical"):  ("Market", 6),
}

# O+O source row offsets relative to each country's "Country (Currency)" anchor row.
# From the pivot table: anchor+7=CERAVE, anchor+8=LRP for MY (2 brands);
#                       anchor+6=CERAVE, anchor+7=LRP, anchor+8=VICHY for SG (3 brands).
# MY only writes LRP (BRAND_07) — no CERAVE row in the suncare split Brand sheet.
OO_BRAND_ROWS: dict[str, list[int]] = {
    "MY": [7],
    "SG": [6, 7, 8],
}

BRAND_SOURCE_MAP = {
    "CERAVE":         "BRAND_06",
    "LA ROCHE POSAY": "BRAND_07",
    "VICHY":          "BRAND_08",
}

OUTPUT_BRAND_MAP = {
    ("Market",   "Mass Medical"):     ("Medic Market",   "Mass Medical"),
    ("Market",   "Non-Mass Medical"): ("Medic Market",   "Non-Mass Medical"),
    ("BRAND_06", None):               ("Cerave",         "Mass Medical"),
    ("BRAND_07", None):               ("La Roche Posay", "Non-Mass Medical"),
    ("BRAND_08", None):               ("Vichy",          "Non-Mass Medical"),
}

FINAL_OUTPUT_ORDER: dict[str, list[tuple[str, str | None]]] = {
    "MY": [
        ("Market",   "Mass Medical"),
        ("Market",   "Non-Mass Medical"),
        ("BRAND_07", None),
    ],
    "SG": [
        ("Market",   "Mass Medical"),
        ("Market",   "Non-Mass Medical"),
        ("BRAND_06", None),
        ("BRAND_07", None),
        ("BRAND_08", None),
    ],
}

MASS_MEDIC = [
    "ACNE AID", "ACNES", "AVEENO", "BALNEUM", "BENZAC", "BIO-OIL", "CARMEX", "CERAVE", "CETAPHIL", "CUREL",
    "DERMATIX", "DERMAVEEN", "DIFFERIN", "DR.G", "DR.YU", "EGO", "EUBOS", "LACTACYD", "LINOLA", "MUSTELA",
    "NEUTROGENA", "PANOXYL", "PHYSIOGEL", "SEBAMED", "TOPICREM", "VANICREAM", "WIS", "XHEKPON", "FIRST AID BEAUTY",
    "AQUAPHOR", "NOBACTER", "LUBRIDERM", "NEOSPORIN", "DARROW", "DEXERYL", "ALERGIBON", "ALPHYGIENE", "BABIGOZ",
    "CANDERMYL", "GALDERMA", "GALDERMA OTHER", "HELIOBLOC", "HYDRODERM OMEGA", "IOCON", "IONIL", "MACROLANE",
    "MICROBAN", "MICROSUN", "NESTLE", "NUTRASPA", "OBSERVANCE", "PHYGIENE", "R-GEN", "SENTIAL", "ACHE", "ACNAID",
    "ACNE FREE", "ACOFAR", "ADDAX", "AKILDIA", "ALBOLENE", "AMLACTIN", "ANSEBIC", "AQUA SOAP", "AQUA-SOAP",
    "AVITIL", "AZULENNE", "BACCIDE", "BEAUTY PLUS", "BEDOOK", "BEPANTHEN/BEPANTHOL", "BETAGRANULOS", "BIAFINE",
    "BIOBLAS", "BIOCLIN", "BIOLIQ", "BIOXCIN", "BLUE LIZARD", "BODYSOL", "BONAVEN", "BOROLINE", "CERAMOL",
    "CERTAIN DRI", "CETOPIC", "CHICCO", "CICAMEL", "COOPER", "COTARYL", "CRISTALIA", "DECUBAL", "DERMAC",
    "DERMACTIVE", "DERMADRATE", "DERMAGE", "DERMAKERI", "DERMENA", "DERMON", "DERSUPRIL", "DEUMAVAN",
    "DOCTISSIMO PARAPHARMACIE", "DR.LI", "DR.LIDERMO", "DRAYEX", "DX2", "E45", "ELDOPAQUE", "EMOLIENTA", "EMOLIN",
    "EMOLIUM", "EPIMAX", "EVASOL", "FARMOQUIMICA", "FILTROSOL", "FLUOCIN", "FREI OEL (BOUHON)", "GALENCO", "GIFRER",
    "GILBERT", "GOLD BOND", "HAMILTON", "HIDRAFIL", "HIPOSOL", "HYALIX", "IDROVEL", "IHADA", "INFASIL",
    "INTERAPOTHEK", "IRALTONE", "ITANIDERM", "KAMILODERM", "KETOXIN", "KINERASE", "KORA", "LACTIBON",
    "LACTO CALAMINE", "LETI", "LIFAR", "LIPODERM", "LOTRIMIN", "MARQUE VERTE", "MICRORET", "MITOSYL", "MODERM",
    "MULTIDERMOL", "MUSSVITAL", "NEUTRA LICE", "NEUTRAPHARM", "NORDIN", "NUMIS", "NUMIS MED", "NURAPHARM",
    "NUTREM", "NUTRISIL", "OILATUM", "OILLAN", "OSMIN", "OTC IBERICA", "PANVEL DERMATIV", "PARABOTICA",
    "PHARMACTIV", "PHARMASEPT", "PHISOHEX", "PROCICAR", "REGENERUM", "RESTIV", "RESTIVOIL", "REVALESKIN", "ROCHE",
    "ROGE CAVAILLES", "ROYALCARE", "RUGARD (SCHEFFLER)", "SALILEX", "SALLVE", "SARNA", "SAUGELLA", "SEBORADIN",
    "SHADE", "SMOOTH-E", "SOLAR FOAM", "S-OLE", "SPECTRABAN", "STANHOME FAMILY EXPERT", "STIEFEL", "STIEPROX",
    "STIPROX", "STIPROXAL", "TARMED", "TRACTOPON", "TRI DERMA MD", "UREADERM", "UVEIL-PS", "UVESOL", "VEA",
    "VENUSIA", "VITA CITRAL", "VITALIFE", "ZODIAC", "QV", "BOBAI", "COLLAGE", "DERMAREST", "EPIZONE E",
    "GLAMY LAB", "LU MILD", "NOLAVER", "OXECURE", "RIUP", "SEBCUR", "SELENGENA", "SEROPIPE", "SHAAN",
    "STAR VILLE", "STRONGVILLE", "SYNOBAR", "UREMOL", "URISEC", "ZINPLEX",
]

NON_MASS_MEDIC_EXTRA = [
    "EUCERIN", "LA ROCHE POSAY", "VICHY", "AVENE", "DR MORITA", "HIRUSCAR", "URIAGE", "SKINCEUTICALS", "DECLEOR",
    "SANOFLORE", "AQUAPHOR", "BIODERMA", "ISDIN", "LIERAC", "FILORGA", "PROACTIV", "ROC", "WINONA", "DR CILABO",
    "EMOLIUM", "PHARMACERIS", "CAUDALIE", "NUXE", "RODAN", "FIELDS", "LIBREDERM", "MANTECORP", "DR. CI : LABO",
]

LOREAL_BRANDS = {
    "LA ROCHE POSAY",
    "VICHY",
    "SKINCEUTICALS",
    "CERAVE",
}

# ---------------------------------------------------------------------------
# Top Brand competitor configuration
# ---------------------------------------------------------------------------

TOP_BRAND_CONFIG = {
    "MY": {"sheet": "1-Malaysia Key Account", "start_row": 53,},
    "SG": {"sheet": "1-Modern Trade","start_row": 54,},
}


def load_non_mass_medic_brands() -> set[str]:
    mapping = pd.read_excel(MAPPING_PATH, sheet_name="Medic")
    from_mapping = mapping.iloc[:, 0].dropna().astype(str).str.upper().str.strip().tolist()
    return set(from_mapping + NON_MASS_MEDIC_EXTRA)


NON_MASS_MEDIC   = load_non_mass_medic_brands()
MASS_MEDIC_SET   = {brand.upper().strip() for brand in MASS_MEDIC}


# ===========================================================================
# Helpers
# ===========================================================================

def parse_my_period(header: str) -> tuple[int, int] | None:
    """
    Converts MY format like:
    'AMJ 24 - w/e 30/06/24' → (2024, 2)
    """

    if not header:
        return None

    text = str(header)

    if "AMJ" in text:
        q = 2
    elif "JAS" in text:
        q = 3
    elif "OND" in text:
        q = 4
    elif "JFM" in text:
        q = 1
    else:
        return None

    try:
        year_part = text.split()[1]  # "24"
        year = 2000 + int(year_part)
    except:
        return None

    return (year, q)

def parse_target_month(value: str) -> tuple[int, int]:
    try:
        year_text, month_text = value.split("-", 1)
        year = int(year_text)
        month = int(month_text)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Use YYYY-MM, for example 2026-04") from exc
    if month < 1 or month > 12:
        raise argparse.ArgumentTypeError("Month must be between 01 and 12")
    return year, month


def month_label(year: int, month: int) -> str:
    return f"{calendar.month_abbr[month]}'{str(year)[-2:]}"


def quote_sheet(name: str) -> str:
    return name.replace("'", "''")


def external_ref(workbook_name: str, sheet_name: str, cell_ref: str) -> str:
    workbook_name = workbook_name.replace("'", "''")
    return f"'[{workbook_name}]{quote_sheet(sheet_name)}'!${cell_ref}"


def external_formula(workbook_name: str, sheet_name: str, cell_ref: str) -> str:
    return f"={external_ref(workbook_name, sheet_name, cell_ref)}"


def excel_col(number: int) -> str:
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result


def split_month_col(sheet_type: str, year: int, month: int) -> str:
    start_col = 4 if sheet_type == "Brand" else 5
    return excel_col(start_col + (year - 2024) * 12 + (month - 1))


def split_source_cell(
    country: str, brand: str, mass_split: str | None, year: int, month: int
) -> tuple[str, str]:
    split_sheet, split_row = SPLIT_SOURCE_ROWS[(country, brand, mass_split)]
    split_col = split_month_col(split_sheet, year, month)
    return split_sheet, f"{split_col}{split_row}"


def split_source_ref(
    split_file_name: str,
    country: str,
    brand: str,
    mass_split: str | None,
    year: int,
    month: int,
) -> str:
    split_sheet, cell_ref = split_source_cell(country, brand, mass_split, year, month)
    return external_ref(split_file_name, split_sheet, cell_ref)


def split_quarter_sum_ref(
    split_file_name: str,
    country: str,
    brand: str,
    mass_split: str | None,
    year: int,
    q_start: int,
) -> str:
    split_sheet, first_cell = split_source_cell(country, brand, mass_split, year, q_start)
    _, last_cell = split_source_cell(country, brand, mass_split, year, q_start + 2)
    workbook_name = split_file_name.replace("'", "''")
    prefix = f"'[{workbook_name}]{quote_sheet(split_sheet)}'!"
    return f"SUM({prefix}${first_cell}:${last_cell})"


def period_key(year: int, month: int) -> int:
    return year * 12 + month


def quarter_fully_within_cutoff(year: int, q_start: int, cutoff_key: int) -> bool:
    return period_key(year, q_start + 2) <= cutoff_key


def quarter_start(month: int) -> int:
    return ((month - 1) // 3) * 3 + 1


def previous_month(year: int, month: int) -> tuple[int, int]:
    if month == 1:
        return year - 1, 12
    return year, month - 1


def capped_actual_period(
    split_max_year: int,
    split_max_month: int,
    target_year: int,
    target_month: int,
) -> tuple[int, int]:
    previous_year, previous_month_number = previous_month(target_year, target_month)
    if period_key(split_max_year, split_max_month) <= period_key(previous_year, previous_month_number):
        return split_max_year, split_max_month
    return previous_year, previous_month_number


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.number_format:
            target.number_format = source.number_format


def get_groups(ws) -> list[tuple[int, int, str, str | None]]:
    groups = []
    start = 2
    for row in range(3, ws.max_row + 2):
        same_group = False
        if row <= ws.max_row:
            same_group = (
                ws.cell(row, 5).value == ws.cell(row - 1, 5).value
                and ws.cell(row, 7).value == ws.cell(row - 1, 7).value
            )
        if not same_group:
            groups.append((start, row - 1, ws.cell(start, 5).value, ws.cell(start, 7).value))
            start = row
    return groups


def build_row_lookup(ws) -> dict[tuple[str, str | None, int, int], int]:
    lookup = {}
    for row in range(2, ws.max_row + 1):
        brand = ws.cell(row, 5).value
        mass_split = ws.cell(row, 7).value
        year = ws.cell(row, 3).value
        month = ws.cell(row, 4).value
        if brand and year and month:
            lookup[(brand, mass_split, int(year), int(month))] = row
    return lookup


# ===========================================================================
# Pulse report discovery
# ===========================================================================

def find_report_rows_and_quarters(
    report_path: Path, country: str
) -> tuple[dict[str, tuple[str, int]], dict[tuple[int, int], str]]:
    wb = load_workbook(report_path, data_only=True, read_only=True)
    targets = BRAND_REPORT_ROWS[country]

    required_sheets = {sheet for sheet, _ in targets.values()}
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"{report_path.name}: missing sheet {sheet_name!r}")

    row_by_brand: dict[str, tuple[str, int]] = dict(targets)

    first_sheet = next(iter(required_sheets))
    ws = wb[first_sheet]
    col_by_quarter: dict[tuple[int, int], str] = {}

    print(f"\n{report_path.name}")
    print(f"Sheet: {first_sheet}")

    for col in range(1, min(ws.max_column, 30) + 1):
        print(
            f"col {col}:",
            ws.cell(REPORT_HEADER_ROW - 1, col).value,
            "|",
            ws.cell(REPORT_HEADER_ROW, col).value,
        )

    col_by_quarter = {}

    for col in range(1, ws.max_column + 1):
        if str(ws.cell(REPORT_HEADER_ROW - 1, col).value or "").strip() != "Sales Value":
            continue

        header = ws.cell(REPORT_HEADER_ROW, col).value
        if not header:
            continue

        if country == "SG":
            if not str(header).startswith("Q"):
                continue
            quarter = int(str(header)[1])
            year = 2000 + int(str(header).split()[1])
            col_by_quarter[(year, quarter)] = excel_col(col)

        else:  # MY
            parsed = parse_my_period(header)
            if parsed:
                col_by_quarter[parsed] = excel_col(col)

    # safety check (IMPORTANT)
    if not row_by_brand:
        raise ValueError(f"{report_path.name}: no brand rows detected")

    if not col_by_quarter:
        raise ValueError(f"{report_path.name}: no quarter columns detected for {country}")

    return row_by_brand, col_by_quarter
# ===========================================================================
# Country block helpers
# ===========================================================================

def find_competitor_rows(
    report_path: Path,
    country: str
):

    wb = load_workbook(report_path, data_only=True, read_only=True)

    config = TOP_BRAND_CONFIG[country]

    ws = wb[config["sheet"]]

    competitors = {}

    for row in range(config["start_row"], ws.max_row + 1):

        brand = ws.cell(row, REPORT_PRODUCT_COL).value

        if not brand:
            continue

        brand = str(brand).strip().upper()

        # Ignore L'Oreal brands
        if brand in LOREAL_BRANDS:
            continue
        competitors[brand] = row

    return competitors

def find_country_blocks(ws) -> dict[str, int]:
    blocks = {}
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value != "Country (Currency)":
            continue
        country_text = str(ws.cell(row, 2).value or "").upper()
        if country_text.startswith("MY "):
            blocks["MY"] = row
        elif country_text.startswith("SG "):
            blocks["SG"] = row
    missing = sorted({"MY", "SG"} - set(blocks))
    if missing:
        raise ValueError(f"{ws.title}: missing country block(s): {', '.join(missing)}")
    return blocks


def extract_period_columns(ws, year_row: int, month_row: int, start_col: int) -> list[tuple[int, int, int]]:
    periods = []
    for col in range(start_col, ws.max_column + 1):
        year = ws.cell(year_row, col).value
        month = ws.cell(month_row, col).value
        if isinstance(year, int) and isinstance(month, int):
            periods.append((col, year, month))
    if not periods:
        raise ValueError(f"{ws.title}: no monthly period columns found")
    return periods


def copy_period_headers(ws, start_col: int, periods: list[tuple[int, int, int]]) -> None:
    for offset, (_, year, month) in enumerate(periods):
        col = start_col + offset
        ws.cell(1, col).value = year
        ws.cell(2, col).value = month


# ===========================================================================
# Build split-method workbook
# ===========================================================================

def build_split_method_workbook(oo_path: Path, output_path: Path) -> tuple[int, int]:
    source_wb = load_workbook(oo_path, data_only=True, read_only=True)
    for sheet_name in ["Market", "L'Oreal "]:
        if sheet_name not in source_wb.sheetnames:
            raise ValueError(f"{oo_path.name}: missing sheet {sheet_name!r}")

    market_source = source_wb["Market"]
    brand_source  = source_wb["L'Oreal "]
    market_blocks = find_country_blocks(market_source)
    brand_blocks  = find_country_blocks(brand_source)

    market_periods = extract_period_columns(
        market_source, market_blocks["MY"] + 4, market_blocks["MY"] + 5, 4
    )
    brand_periods = extract_period_columns(
        brand_source, brand_blocks["MY"] + 4, brand_blocks["MY"] + 5, 3
    )
    # Global max period — driven by whichever brand has the most data.
    # VICHY's shorter range is handled per-brand inside calculate_sellout_values.
    max_year, max_month = max(
        (year, month) for _, year, month in brand_periods + market_periods
    )

    output_wb = Workbook()
    market_ws  = output_wb.active
    market_ws.title = "Market"
    brand_ws   = output_wb.create_sheet("Brand")

    for col, value in enumerate(["Country", "O+O Brand", "Mass/Non-Mass", "Category"], start=1):
        market_ws.cell(2, col).value = value
    for col, value in enumerate(["Country", "O+O Brand", "Category"], start=1):
        brand_ws.cell(2, col).value = value

    copy_period_headers(market_ws, 5, market_periods)
    copy_period_headers(brand_ws,  4, brand_periods)

    market_output_row = 3
    for country in ["MY", "SG"]:
        block = market_blocks[country]
        for source_row in [block + 6, block + 7]:
            market_ws.cell(market_output_row, 1).value = country
            market_ws.cell(market_output_row, 2).value = market_source.cell(source_row, 1).value
            market_ws.cell(market_output_row, 3).value = market_source.cell(source_row, 2).value
            market_ws.cell(market_output_row, 4).value = market_source.cell(source_row, 3).value
            for offset, (source_col, _, _) in enumerate(market_periods):
                market_ws.cell(market_output_row, 5 + offset).value = (
                    market_source.cell(source_row, source_col).value
                )
            market_output_row += 1

    brand_output_row = 3
    for country in ["MY", "SG"]:
        block   = brand_blocks[country]
        offsets = OO_BRAND_ROWS[country]
        for offset in offsets:
            source_row   = block + offset
            raw_brand    = str(brand_source.cell(source_row, 1).value or "").strip()
            output_brand = BRAND_SOURCE_MAP.get(raw_brand.upper(), raw_brand)
            category     = brand_source.cell(source_row, 2).value
            brand_ws.cell(brand_output_row, 1).value = country
            brand_ws.cell(brand_output_row, 2).value = output_brand
            brand_ws.cell(brand_output_row, 3).value = category
            for col_offset, (source_col, _, _) in enumerate(brand_periods):
                brand_ws.cell(brand_output_row, 4 + col_offset).value = (
                    brand_source.cell(source_row, source_col).value
                )
            brand_output_row += 1

    for ws in [market_ws, brand_ws]:
        header_fill = PatternFill("solid", fgColor="D9EAD3")
        for row in [1, 2]:
            for cell in ws[row]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
        for column_cells in ws.columns:
            ws.column_dimensions[column_cells[0].column_letter].width = 14

    output_wb.save(output_path)
    return max_year, max_month


# ===========================================================================
# Export picked O+O values to CSV
# ===========================================================================

def export_oo_picked_values(oo_path: Path, output_path: Path) -> int:
    source_wb = load_workbook(oo_path, data_only=True, read_only=True)
    rows = []

    brand_source = source_wb["L'Oreal "]
    brand_blocks = find_country_blocks(brand_source)
    for country in ["MY", "SG"]:
        block   = brand_blocks[country]
        periods = extract_period_columns(brand_source, block + 4, block + 5, 3)
        for offset in OO_BRAND_ROWS[country]:
            source_row   = block + offset
            raw_brand    = str(brand_source.cell(source_row, 1).value or "").strip()
            output_brand = BRAND_SOURCE_MAP.get(raw_brand.upper(), raw_brand)
            category     = brand_source.cell(source_row, 2).value
            for source_col, year, month in periods:
                rows.append({
                    "Output Sheet":  "Brand",
                    "Country":       country,
                    "Source Sheet":  "L'Oreal ",
                    "Source Cell":   brand_source.cell(source_row, source_col).coordinate,
                    "Source Brand":  raw_brand,
                    "Output Brand":  output_brand,
                    "Mass/Non-Mass": "",
                    "Category":      category,
                    "Year":          year,
                    "Month":         month,
                    "Value":         brand_source.cell(source_row, source_col).value,
                })

    market_source = source_wb["Market"]
    market_blocks = find_country_blocks(market_source)
    for country in ["MY", "SG"]:
        block   = market_blocks[country]
        periods = extract_period_columns(market_source, block + 4, block + 5, 4)
        for source_row in [block + 6, block + 7]:
            brand      = market_source.cell(source_row, 1).value
            mass_split = market_source.cell(source_row, 2).value
            category   = market_source.cell(source_row, 3).value
            for source_col, year, month in periods:
                rows.append({
                    "Output Sheet":  "Market",
                    "Country":       country,
                    "Source Sheet":  "Market",
                    "Source Cell":   market_source.cell(source_row, source_col).coordinate,
                    "Source Brand":  brand,
                    "Output Brand":  brand,
                    "Mass/Non-Mass": mass_split,
                    "Category":      category,
                    "Year":          year,
                    "Month":         month,
                    "Value":         market_source.cell(source_row, source_col).value,
                })

    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)

def get_competitor_mass_split(
    brand: str,
    missing_brands: set
):

    brand = brand.upper().strip()


    if brand in MASS_MEDIC_SET:
        return "Mass Medical"


    if brand in NON_MASS_MEDIC_EXTRA:
        return "Non-Mass Medical"


    missing_brands.add(brand)

    return "Non-Mass Medical"

def split_competitor_quarter(quarter_value, wob_values):

    total = sum(wob_values)


    if total == 0:
        return [
            0,
            0,
            0
        ]

    return [
        quarter_value * x / total
        for x in wob_values
    ]

def calculate_competitor_values(
    report_path: Path,
    split_path: Path,
    country: str,
    target_year: int,
    target_month: int,
    report_quarters: dict[tuple[int, int], str],
):
    """Calculate competitor values including estimation for future quarters."""
    
    report_wb = load_workbook(report_path, data_only=True, read_only=True)
    split_wb = load_workbook(split_path, data_only=True, read_only=True)
    competitor_rows = find_competitor_rows(report_path, country)

    missing_brands = set()
    output = []

    sheet = TOP_BRAND_CONFIG[country]["sheet"]
    ws = report_wb[sheet]

    # Get all quarters from the report
    available_quarters = sorted(report_quarters.keys())
    if not available_quarters:
        print(f"WARNING: No quarters found for {country}")
        return output

    # For each competitor brand
    for brand, row in competitor_rows.items():
        mass_split = get_competitor_mass_split(brand, missing_brands)
        
        # Determine which quarters to process
        # Process all available quarters plus estimate future quarters up to target
        quarters_to_process = []
        
        # Add all available quarters from report
        quarters_to_process.extend(available_quarters)
        
        # Add future quarters up to target
        last_year, last_q = available_quarters[-1]
        
        # Calculate target quarter
        target_q = (target_month - 1) // 3 + 1
        
        # Generate future quarters
        year = last_year
        q = last_q + 1
        
        while year < target_year or (year == target_year and q <= target_q):
            quarters_to_process.append((year, q))
            q += 1
            if q > 4:
                q = 1
                year += 1
        
        # Process each quarter
        for year, quarter in quarters_to_process:
            q_start = (quarter - 1) * 3 + 1
            
            # Check if this quarter is available in the pulse report
            is_available = (year, quarter) in report_quarters
            
            if is_available:
                # Get quarter value from pulse report
                col_letter = report_quarters[(year, quarter)]
                col = column_index_from_string(col_letter)
                quarter_value = ws.cell(row, col).value
                
                if quarter_value is None or quarter_value == 0:
                    continue
            else:
                # For future quarters, use previous quarter's values
                # Get previous quarter
                if quarter == 1:
                    prev_year = year - 1
                    prev_q = 4
                else:
                    prev_year = year
                    prev_q = quarter - 1
                
                # Check if previous quarter is available
                if (prev_year, prev_q) not in report_quarters:
                    # If previous quarter isn't available, skip
                    continue
                
                # Get the previous quarter's monthly values from output cache
                # We need to calculate this brand's previous quarter values first
                # Since we're processing in order, the previous quarter should have been processed
                # But for the first quarter after available data, we need to get it from the report
                
                # For simplicity, we'll compute the previous quarter values by looking at the report
                # if the previous quarter is available, or by using the previous quarter's previous quarter
                prev_col_letter = report_quarters[(prev_year, prev_q)]
                prev_col = column_index_from_string(prev_col_letter)
                prev_q_value = ws.cell(row, prev_col).value
                
                if prev_q_value is None or prev_q_value == 0:
                    continue
                
                # For estimation, we need the monthly split of the previous quarter
                # Get the WOB for the previous quarter
                wob = []
                for month in range(1, 4):
                    wob.append(read_split_value(
                        split_wb, country, "Market", mass_split, 
                        prev_year, (prev_q - 1) * 3 + month
                    ))
                
                # Split the previous quarter value using market WOB
                monthly_values = split_competitor_quarter(prev_q_value, wob)
                
                # For current quarter, use the same monthly values
                # (copy previous quarter's months to current quarter)
                for i, value in enumerate(monthly_values):
                    current_month = q_start + i
                    
                    # Only include months up to target_month
                    if year == target_year and current_month > target_month:
                        continue
                    
                    output.append({
                        "Country": country,
                        "Brand": brand,
                        "Year": year,
                        "Month": current_month,
                        "Value": value,
                        "Mass/Non-Mass": mass_split,
                        "IsEstimated": True
                    })
                
                continue  # Skip the rest of the loop for estimated quarters
            
            # For available quarters, get WOB and split
            wob = []
            for month in range(1, 4):
                wob.append(read_split_value(
                    split_wb, country, "Market", mass_split, 
                    year, q_start + month - 1
                ))
            
            monthly_values = split_competitor_quarter(quarter_value, wob)
            
            for i, value in enumerate(monthly_values):
                current_month = q_start + i
                
                # Only include months up to target_month
                if year == target_year and current_month > target_month:
                    continue
                
                output.append({
                    "Country": country,
                    "Brand": brand,
                    "Year": year,
                    "Month": current_month,
                    "Value": value,
                    "Mass/Non-Mass": mass_split,
                    "IsEstimated": False
                })

    if missing_brands:
        print("\nMissing competitor classification:")
        for b in sorted(missing_brands):
            print(f"  {b} => Using Non-Mass Medical WOB")

    return output

# ===========================================================================
# Read helpers
# ===========================================================================

def read_split_value(
    split_wb, country: str, brand: str, mass_split: str | None, year: int, month: int
) -> float:
    sheet_name, cell_ref = split_source_cell(country, brand, mass_split, year, month)
    value = split_wb[sheet_name][cell_ref].value
    if value is None:
        raise ValueError(
            f"Missing split value for {country} {brand} {mass_split or ''} {year}-{month:02d}"
        )
    return float(value)


def try_read_split_value(
    split_wb, country: str, brand: str, mass_split: str | None, year: int, month: int
) -> float | None:
    """Like read_split_value but returns None instead of raising for missing/zero data."""
    key = (country, brand, mass_split)
    if key not in SPLIT_SOURCE_ROWS:
        return None
    try:
        v = read_split_value(split_wb, country, brand, mass_split, year, month)
        return v if v != 0.0 else None
    except (ValueError, KeyError):
        return None


def quarter_split_total(
    split_wb, country: str, brand: str, mass_split: str | None,
    year: int, q_start: int
) -> float | None:
    """Sum of split values for a full quarter. Returns None if any month is missing/zero."""
    total = 0.0
    for m in range(q_start, q_start + 3):
        v = try_read_split_value(split_wb, country, brand, mass_split, year, m)
        if v is None:
            return None
        total += v
    return total if total > 0.0 else None

def quarter_has_missing_brand_data(
    split_wb,
    country: str,
    brand: str,
    year: int,
    q_start: int,
) -> bool:
    """
    Returns True if ANY month in the quarter is missing/zero.
    """
    for m in range(q_start, q_start + 3):
        v = try_read_split_value(
            split_wb,
            country,
            brand,
            None,
            year,
            m,
        )
        if v is None:
            return True

    return False

def read_report_value(
    report_wb,
    report_rows: dict[str, tuple[str, int]],
    report_quarters: dict[tuple[int, int], str],
    brand: str,
    year: int,
    quarter: int,
    country: str = ""
) -> float:

    sheet_name, row = report_rows[brand]

    col_letter = report_quarters.get((year, quarter))
    if not col_letter:
        print(f"WARNING: Missing quarter mapping -> {country} | {brand} | {year} Q{quarter} -> treated as 0")
        return 0.0

    value = report_wb[sheet_name].cell(
        row,
        column_index_from_string(col_letter)
    ).value

    if value is None or value == "":
        print(f"WARNING: Missing pulse value -> {country} | {brand} | {year} Q{quarter} -> treated as 0")
        return 0.0

    return float(value)


def read_split_max_period(split_path: Path) -> tuple[int, int]:
    """Return the latest period header in the split file.

    The global split_max is driven by the brands that have the most data
    (CERAVE, LRP, Market).  VICHY's shorter data range is handled separately
    inside calculate_sellout_values via per-brand cutoff logic.
    """
    wb = load_workbook(split_path, data_only=True, read_only=True)
    periods = []
    for sheet_name, start_col in [("Brand", 4), ("Market", 5)]:
        ws = wb[sheet_name]
        for col in range(start_col, ws.max_column + 1):
            year  = ws.cell(1, col).value
            month = ws.cell(2, col).value
            if isinstance(year, int) and isinstance(month, int):
                periods.append((year, month))
    if not periods:
        raise ValueError(f"{split_path.name}: no monthly periods found")
    return max(periods)


# ===========================================================================
# Safe fallback helper
# ---------------------------------------------------------------------------
# Returns the K value for a given row key, or None if the row doesn't exist
# or its K value hasn't been set yet.  Avoids KeyError on None keys.
# ===========================================================================

def _safe_k(k_values: dict, row_lookup: dict, brand, mass_split, year, q_start) -> float | None:
    key = (brand, mass_split, year, q_start)
    row = row_lookup.get(key)
    if row is None:
        return None
    return k_values.get(row)


# ===========================================================================
# Calculate sellout values
# ===========================================================================

def _brand08_split_max(split_wb) -> tuple[int, int] | None:
    """Find BRAND_08's own last month with actual O+O data in the split Brand sheet.

    Row 6 of the Brand sheet is SG BRAND_08 (Vichy). We scan its data columns
    and return the last month where the cell is non-None and non-zero.
    Returns None if no data found at all.
    """
    ws = split_wb["Brand"]
    last = None
    for col in range(4, ws.max_column + 1):
        year  = ws.cell(1, col).value
        month = ws.cell(2, col).value
        if not (isinstance(year, int) and isinstance(month, int)):
            continue
        val = ws.cell(6, col).value  # row 6 = SG BRAND_08
        if val is not None and float(val) != 0.0:
            last = (year, month)
    return last

def calculate_sellout_values(
    estimation_wb,
    split_path: Path,
    report_paths: dict[str, Path],
    report_info: dict[str, tuple[dict[str, tuple[str, int]], dict[tuple[int, int], str]]],
    split_max_year: int,
    split_max_month: int,
    target_year: int,
    target_month: int,
) -> dict[tuple[str, int], float]:
    """Calculate the final L (sellout estimate) for every row in MY/SG LDB sheets."""
    
    split_wb = load_workbook(split_path, data_only=True, read_only=True)
    report_wbs = {
        country: load_workbook(path, data_only=True, read_only=True)
        for country, path in report_paths.items()
    }

    actual_max_year, actual_max_month = capped_actual_period(
        split_max_year, split_max_month, target_year, target_month
    )
    split_max_key = period_key(actual_max_year, actual_max_month)

    # Per-brand cutoff for BRAND_08 — find its own last populated month
    b08_last = _brand08_split_max(split_wb)
    b08_max_key = period_key(*b08_last) if b08_last else 0

    calculated: dict[tuple[str, int], float] = {}

    for sheet_name in ["MY LDB", "SG LDB"]:
        ws = estimation_wb[sheet_name]
        country = sheet_name[:2]
        report_rows, report_quarters = report_info[country]
        row_lookup = build_row_lookup(ws)
        groups = get_groups(ws)

        h_values: dict[int, float] = {}
        i_values: dict[int, float | None] = {}
        j_values: dict[int, float] = {}
        k_values: dict[int, float | None] = {}
        l_values: dict[int, float] = {}

        # First pass: calculate H values
        for start, end, brand, mass_split in groups:
            brand_has_pulse = brand in report_rows

            for row in range(start, end + 1):
                year = int(ws.cell(row, 3).value)
                month = int(ws.cell(row, 4).value)
                q_start = quarter_start(month)
                quarter = (month - 1) // 3 + 1

                # Determine if this is an estimate row
                if brand == "BRAND_08":
                    is_estimate_row = period_key(year, month) > b08_max_key
                else:
                    is_estimate_row = period_key(year, month) > split_max_key

                # Check if quarter has pulse data
                has_pulse = brand_has_pulse and (year, quarter) in report_quarters

                # For BRAND_08 special handling
                use_cerave_wob_quarter = False
                if brand == "BRAND_08":
                    use_cerave_wob_quarter = quarter_has_missing_brand_data(
                        split_wb, country, "BRAND_08", year, q_start
                    )

                # BRAND_08 special case
                if brand == "BRAND_08" and (is_estimate_row or use_cerave_wob_quarter):
                    pulse_q_total: float | None = None
                    if has_pulse:
                        try:
                            pulse_q_total = read_report_value(
                                report_wbs[country], report_rows, report_quarters,
                                brand, year, quarter, country
                            )
                            if pulse_q_total == 0.0:
                                pulse_q_total = None
                        except (ValueError, KeyError):
                            pulse_q_total = None

                    if pulse_q_total is None:
                        h_values[row] = 0.0
                        i_values[row] = None
                        j_values[row] = 0.0
                        k_values[row] = None
                        l_values[row] = 0.0
                        continue

                    # Use Cerave WOB for BRAND_08 estimation
                    b06_month_val = try_read_split_value(
                        split_wb, country, "BRAND_06", None, year, month
                    )
                    b06_q_total = quarter_split_total(
                        split_wb, country, "BRAND_06", None, year, q_start
                    )
                    if b06_q_total and b06_q_total > 0.0 and b06_month_val is not None:
                        weight = b06_month_val / b06_q_total
                    else:
                        weight = 1.0 / 3.0

                    h_values[row] = 0.0
                    i_values[row] = None
                    j_values[row] = weight
                    k_values[row] = pulse_q_total if month == q_start else None
                    l_values[row] = pulse_q_total * weight
                    continue

                # ---- H ----
                if is_estimate_row:
                    # For estimation, get the same month from the previous quarter
                    prev_q_year = year
                    prev_q_month = month - 3
                    if prev_q_month <= 0:
                        prev_q_month += 12
                        prev_q_year -= 1
                    
                    # Get the row for the same month in the previous quarter
                    previous_q_key = (brand, mass_split, prev_q_year, prev_q_month)
                    previous_q_row = row_lookup.get(previous_q_key)
                    
                    if previous_q_row is not None:
                        # Check if we already have h_values for this row
                        if previous_q_row in h_values:
                            h_values[row] = h_values[previous_q_row]
                        else:
                            # If h_values not available, try to read from split workbook
                            try:
                                h_values[row] = read_split_value(
                                    split_wb, country, brand, mass_split, prev_q_year, prev_q_month
                                )
                            except (ValueError, KeyError):
                                h_values[row] = 0.0
                    else:
                        h_values[row] = 0.0
                else:
                    # Actual data - read from split workbook
                    try:
                        h_values[row] = read_split_value(
                            split_wb, country, brand, mass_split, year, month
                        )
                    except (ValueError, KeyError):
                        h_values[row] = 0.0

                # ---- I ----
                i_values[row] = None if row == start else h_values[row - 1] / h_values[row] - 1 if h_values[row] != 0 else 0

        # Second pass: calculate J, K, L values
        for start, end, brand, mass_split in groups:
            brand_has_pulse = brand in report_rows

            for row in range(start, end + 1):
                year = int(ws.cell(row, 3).value)
                month = int(ws.cell(row, 4).value)
                q_start = quarter_start(month)
                quarter = (month - 1) // 3 + 1

                if brand == "BRAND_08":
                    is_estimate_row = period_key(year, month) > b08_max_key
                else:
                    is_estimate_row = period_key(year, month) > split_max_key

                use_cerave_wob_quarter = False
                if brand == "BRAND_08":
                    use_cerave_wob_quarter = quarter_has_missing_brand_data(
                        split_wb, country, "BRAND_08", year, q_start
                    )

                has_pulse = brand_has_pulse and (year, quarter) in report_quarters

                if brand == "BRAND_08" and (is_estimate_row or use_cerave_wob_quarter):
                    pulse_q_total: float | None = None
                    if has_pulse:
                        try:
                            pulse_q_total = read_report_value(
                                report_wbs[country], report_rows, report_quarters,
                                brand, year, quarter, country
                            )
                            if pulse_q_total == 0.0:
                                pulse_q_total = None
                        except (ValueError, KeyError):
                            pulse_q_total = None

                    if pulse_q_total is None:
                        j_values[row] = 0.0
                        k_values[row] = None
                        l_values[row] = 0.0
                        continue

                    # Skip BRAND_08 estimation (already handled in first pass)
                    continue

                # ---- J ----
                denominator_year = year - 1 if is_estimate_row else year
                denominator_q_start = q_start
                denominator_in_split = quarter_fully_within_cutoff(
                    denominator_year, denominator_q_start, split_max_key
                )

                if brand == "Market":
                    if denominator_in_split:
                        denominator = sum(
                            read_split_value(split_wb, country, "Market", split, denominator_year, q_month)
                            for split in ["Mass Medical", "Non-Mass Medical"]
                            for q_month in range(denominator_q_start, denominator_q_start + 3)
                        )
                    else:
                        denominator = sum(
                            h_values[r]
                            for split in ["Mass Medical", "Non-Mass Medical"]
                            for r in quarter_rows(
                                row_lookup, "Market", split, denominator_year, denominator_q_start
                            )
                        )
                    j_values[row] = h_values[row] / denominator if denominator != 0 else 0.0

                    # ---- K (Market) ----
                    should_hold_actual = (
                        not is_estimate_row
                        and has_pulse
                        and mass_split == "Mass Medical"
                        and month == q_start
                    )
                    if should_hold_actual:
                        k_values[row] = read_report_value(
                            report_wbs[country], report_rows, report_quarters, brand, year, quarter
                        )
                    else:
                        k_values[row] = None

                    # ---- L (Market) ----
                    k_row_current = row_lookup.get(("Market", "Mass Medical", year, q_start))
                    k_val_current = k_values.get(k_row_current) if k_row_current is not None else None
                    prev_period = previous_month(year, month)
                    previous_row = row_lookup.get((brand, mass_split, prev_period[0], prev_period[1]))

                    if not is_estimate_row and has_pulse and k_val_current is not None:
                        l_values[row] = k_val_current * j_values[row]
                    elif previous_row is not None and previous_row in l_values:
                        l_values[row] = l_values[previous_row] / (1 + i_values[row]) if i_values[row] != -1 else 0.0
                    else:
                        # For estimated rows, use previous quarter's same month value
                        # We need to get the value from h_values and apply the proper scaling
                        prev_q_year = year
                        prev_q_month = month - 3
                        if prev_q_month <= 0:
                            prev_q_month += 12
                            prev_q_year -= 1
                        
                        prev_q_key = (brand, mass_split, prev_q_year, prev_q_month)
                        prev_q_row = row_lookup.get(prev_q_key)
                        
                        if prev_q_row is not None:
                            # Check if we have l_values for the previous quarter's same month
                            if prev_q_row in l_values:
                                l_values[row] = l_values[prev_q_row]
                            elif prev_q_row in h_values:
                                # If we have h_values but not l_values, use h_values * j_values
                                # But we need j_values for the previous quarter's same month
                                # For simplicity, just use h_values
                                l_values[row] = h_values[prev_q_row]
                            else:
                                l_values[row] = 0.0
                        else:
                            l_values[row] = 0.0

                else:
                    # ---- Brand row ----
                    if brand == "BRAND_08":
                        denom_in_split = quarter_fully_within_cutoff(
                            denominator_year, denominator_q_start, b08_max_key
                        )
                    else:
                        denom_in_split = denominator_in_split

                    if denom_in_split:
                        denominator = sum(
                            read_split_value(split_wb, country, brand, mass_split, denominator_year, q_month)
                            for q_month in range(denominator_q_start, denominator_q_start + 3)
                        )
                    else:
                        denominator = sum(
                            h_values[r]
                            for r in quarter_rows(
                                row_lookup, brand, mass_split, denominator_year, denominator_q_start
                            )
                        )
                    j_values[row] = h_values[row] / denominator if denominator else 0.0

                    # ---- K (Brand) ----
                    should_hold_actual = not is_estimate_row and has_pulse and month == q_start
                    if should_hold_actual:
                        k_values[row] = read_report_value(
                            report_wbs[country], report_rows, report_quarters, brand, year, quarter
                        )
                    else:
                        k_values[row] = None

                    # ---- L (Brand) ----
                    k_row_current = row_lookup.get((brand, mass_split, year, q_start))
                    k_val_current = k_values.get(k_row_current) if k_row_current is not None else None
                    prev_period = previous_month(year, month)
                    previous_row = row_lookup.get((brand, mass_split, prev_period[0], prev_period[1]))

                    if not is_estimate_row and has_pulse and k_val_current is not None:
                        l_values[row] = k_val_current * j_values[row]
                    elif previous_row is not None and previous_row in l_values:
                        l_values[row] = l_values[previous_row] / (1 + i_values[row]) if i_values[row] != -1 else 0.0
                    else:
                        # For estimated rows, use previous quarter's same month value
                        prev_q_year = year
                        prev_q_month = month - 3
                        if prev_q_month <= 0:
                            prev_q_month += 12
                            prev_q_year -= 1
                        
                        prev_q_key = (brand, mass_split, prev_q_year, prev_q_month)
                        prev_q_row = row_lookup.get(prev_q_key)
                        
                        if prev_q_row is not None:
                            # Check if we have l_values for the previous quarter's same month
                            if prev_q_row in l_values:
                                l_values[row] = l_values[prev_q_row]
                            elif prev_q_row in h_values:
                                # If we have h_values but not l_values, use h_values * j_values
                                l_values[row] = h_values[prev_q_row] * j_values[row]
                            else:
                                l_values[row] = 0.0
                        else:
                            l_values[row] = 0.0

        for row, value in l_values.items():
            calculated[(sheet_name, row)] = value

    return calculated

# ===========================================================================
# Export final flat output
# ===========================================================================

def export_final_output(
    estimation_wb,
    template_path: Path,
    output_path: Path,
    start_year: int,
    target_year: int,
    target_month: int,
    sellout_values: dict[tuple[str, int], float],
    my_top: list,
    sg_top: list,
) -> int:
    if not template_path.exists():
        raise FileNotFoundError(f"Output template not found: {template_path}")

    output_wb   = load_workbook(template_path)
    template_ws = output_wb.active

    header_values = [
        template_ws.cell(1, col).value
        for col in range(1, template_ws.max_column + 1)
    ]
    header_styles = []
    for col in range(1, template_ws.max_column + 1):
        source_cell = template_ws.cell(1, col)
        header_styles.append({
            "style":         copy(source_cell._style),
            "font":          copy(source_cell.font),
            "fill":          copy(source_cell.fill),
            "border":        copy(source_cell.border),
            "alignment":     copy(source_cell.alignment),
            "number_format": source_cell.number_format,
            "width":         template_ws.column_dimensions[source_cell.column_letter].width,
        })

    for sheet_name in list(output_wb.sheetnames):
        del output_wb[sheet_name]

    output_sheets: dict[str, object] = {}
    for sheet_name in ["MY_SC", "SG_SC", "MY_SC_Top Brands", "SG_SC_Top Brands"]:
        ws = output_wb.create_sheet(sheet_name)
        output_sheets[sheet_name] = ws
        for col, value in enumerate(header_values, start=1):
            cell  = ws.cell(1, col)
            style = header_styles[col - 1]
            cell.value         = value
            cell._style        = copy(style["style"])
            cell.font          = copy(style["font"])
            cell.fill          = copy(style["fill"])
            cell.border        = copy(style["border"])
            cell.alignment     = copy(style["alignment"])
            cell.number_format = style["number_format"]
            ws.column_dimensions[cell.column_letter].width = style["width"]

    row_lookups = {
        sheet_name: build_row_lookup(estimation_wb[sheet_name])
        for sheet_name in ["MY LDB", "SG LDB"]
    }
    output_rows = {"MY": 2, "SG": 2, "MY_TOP": 2, "SG_TOP": 2}

    for year in range(start_year, target_year + 1):
        month_end = target_month if year == target_year else 12
        for month in range(1, month_end + 1):
            for sheet_name in ["MY LDB", "SG LDB"]:
                country = sheet_name[:2]

                if country == "MY":
                    ws = output_sheets["MY_SC"]
                else:
                    ws = output_sheets["SG_SC"]

                output_row = output_rows[country]
                row_lookup = row_lookups[sheet_name]

                for brand, mass_split in FINAL_OUTPUT_ORDER[country]:
                    source_row = row_lookup.get((brand, mass_split, year, month))
                    if source_row is None:
                        raise ValueError(
                            f"{sheet_name}: missing final output source row for "
                            f"{brand} {mass_split or ''} {year}-{month:02d}"
                        )
                    output_brand, output_mass_split = OUTPUT_BRAND_MAP[(brand, mass_split)]
                    ws.cell(output_row,  1).value = country
                    ws.cell(output_row,  2).value = "Offline"
                    ws.cell(output_row,  3).value = year
                    ws.cell(output_row,  4).value = "Offline_Est"
                    ws.cell(output_row,  5).value = "SKINCARE"
                    ws.cell(output_row,  6).value = "Sun Care"
                    ws.cell(output_row,  7).value = output_mass_split
                    ws.cell(output_row,  8).value = output_brand
                    ws.cell(output_row,  9).value = month
                    ws.cell(output_row, 10).value = sellout_values[(sheet_name, source_row)]
                    ws.cell(output_row, 10).number_format = "#,##0.00"
                    ws.cell(output_row, 11).value = None
                    output_row += 1

                output_rows[country] = output_row
    
    # ==============================
    # Write competitor top brands
    # ==============================

    for country, competitor_values in [
        ("MY", my_top),
        ("SG", sg_top)
    ]:

        sheet = f"{country}_SC_Top Brands"

        ws = output_sheets[sheet]

        row = output_rows[f"{country}_TOP"]

        for item in competitor_values:

            ws.cell(row, 1).value = country
            ws.cell(row, 2).value = "Offline"
            ws.cell(row, 3).value = item["Year"]
            ws.cell(row, 4).value = "Offline_Est"
            ws.cell(row, 5).value = "SKINCARE"
            ws.cell(row, 6).value = "Sun Care"
            ws.cell(row, 7).value = item["Mass/Non-Mass"]
            ws.cell(row, 8).value = item["Brand"]
            ws.cell(row, 9).value = item["Month"]
            ws.cell(row,10).value = item["Value"]
            ws.cell(row,10).number_format = "#,##0.00"

            row += 1

        output_rows[f"{country}_TOP"] = row

    output_wb.save(output_path)
    return sum(row - 2 for row in output_rows.values())


# ===========================================================================
# Estimation workbook formula writer
# ===========================================================================

def quarter_rows(row_lookup, brand, mass_split, year, q_start) -> list[int]:
    return [
        row_lookup[(brand, mass_split, year, month)]
        for month in range(q_start, q_start + 3)
        if (brand, mass_split, year, month) in row_lookup
    ]


def sum_range(rows: list[int]) -> str:
    if not rows:
        return "0"
    return f"SUM(H{min(rows)}:H{max(rows)})"


def ensure_target_rows(ws, target_year: int, target_month: int) -> None:
    groups     = get_groups(ws)
    row_lookup = build_row_lookup(ws)
    inserts    = []
    for start, end, brand, mass_split in groups:
        if (brand, mass_split, target_year, target_month) in row_lookup:
            continue
        prev_year, prev_month = previous_month(target_year, target_month)
        previous_row = row_lookup.get((brand, mass_split, prev_year, prev_month))
        if previous_row is None:
            raise ValueError(
                f"{ws.title}: cannot add {target_year}-{target_month:02d} for "
                f"{brand} {mass_split or ''}; previous month is missing"
            )
        if previous_row != end:
            raise ValueError(
                f"{ws.title}: {brand} {mass_split or ''} has rows after the previous month. "
                "Run missing months in order."
            )
        inserts.append(previous_row)
    for previous_row in sorted(inserts, reverse=True):
        ws.insert_rows(previous_row + 1, 1)
        copy_row_style(ws, previous_row, previous_row + 1)
        for col in [1, 2, 5, 6, 7, 13]:
            ws.cell(previous_row + 1, col).value = ws.cell(previous_row, col).value
        ws.cell(previous_row + 1, 3).value = target_year
        ws.cell(previous_row + 1, 4).value = target_month


def update_formulas(
    ws,
    target_year: int,
    target_month: int,
    split_max_year: int,
    split_max_month: int,
    split_file_name: str,
    report_file_name: str,
    report_rows: dict[str, tuple[str, int]],
    report_quarters: dict[tuple[int, int], str],
) -> None:
    country = ws.title[:2]
    ensure_target_rows(ws, target_year, target_month)
    row_lookup = build_row_lookup(ws)
    groups     = get_groups(ws)

    actual_max_year, actual_max_month = capped_actual_period(
        split_max_year, split_max_month, target_year, target_month
    )
    split_max_key = period_key(actual_max_year, actual_max_month)

    for start, end, brand, mass_split in groups:
        brand_has_pulse = brand in report_rows

        for row in range(start, end + 1):
            year    = int(ws.cell(row, 3).value)
            month   = int(ws.cell(row, 4).value)
            q_start = quarter_start(month)
            quarter = (month - 1) // 3 + 1
            is_estimate_row = period_key(year, month) > split_max_key
            has_pulse = brand_has_pulse and (year, quarter) in report_quarters

            report_sheet = report_rows[brand][0] if brand_has_pulse else None
            report_row   = report_rows[brand][1] if brand_has_pulse else None
            ws.cell(row, 13).value = report_sheet

            # ---- H ----
            if is_estimate_row:
                prev_year, prev_month = previous_month(year, month)
                previous_row   = row_lookup[(brand, mass_split, prev_year, prev_month)]
                prior_year_row = row_lookup[(brand, mass_split, year - 1, month)]
                
                prev_q_year = year
                prev_q_month = month - 3

                if prev_q_month <= 0:
                    prev_q_month += 12
                    prev_q_year -= 1


                previous_q_row = row_lookup[
                    (
                        brand,
                        mass_split,
                        prev_q_year,
                        prev_q_month
                    )
                ]


                ws.cell(row, 8).value = f"=H{previous_q_row}"
            else:
                split_sheet, cell_ref = split_source_cell(country, brand, mass_split, year, month)
                ws.cell(row, 8).value = external_formula(split_file_name, split_sheet, cell_ref)

            # ---- I ----
            ws.cell(row, 9).value = None if row == start else f"=H{row - 1}/H{row}-1"

            denominator_year    = year - 1 if is_estimate_row else year
            denominator_q_start = q_start
            denominator_in_split = quarter_fully_within_cutoff(
                denominator_year, denominator_q_start, split_max_key
            )

            if brand == "Market":
                if denominator_in_split:
                    mass_denom     = split_quarter_sum_ref(
                        split_file_name, country, "Market", "Mass Medical",
                        denominator_year, denominator_q_start
                    )
                    non_mass_denom = split_quarter_sum_ref(
                        split_file_name, country, "Market", "Non-Mass Medical",
                        denominator_year, denominator_q_start
                    )
                    denominator = f"{mass_denom}+{non_mass_denom}"
                else:
                    mass_rows     = quarter_rows(
                        row_lookup, "Market", "Mass Medical",
                        denominator_year, denominator_q_start
                    )
                    non_mass_rows = quarter_rows(
                        row_lookup, "Market", "Non-Mass Medical",
                        denominator_year, denominator_q_start
                    )
                    denominator = f"{sum_range(mass_rows)}+{sum_range(non_mass_rows)}"
                ws.cell(row, 10).value = f"=H{row}/({denominator})"

                k_row_current  = row_lookup.get(("Market", "Mass Medical", year, q_start))
                k_row_fallback = row_lookup.get(("Market", "Mass Medical", year - 1, q_start))
                should_hold_actual = (
                    not is_estimate_row
                    and has_pulse
                    and mass_split == "Mass Medical"
                    and month == q_start
                )
                if should_hold_actual:
                    col = report_quarters[(year, quarter)]
                    ws.cell(row, 11).value = external_formula(
                        report_file_name, report_sheet, f"{col}{report_row}"
                    )
                else:
                    ws.cell(row, 11).value = None

                prev_period  = previous_month(year, month)
                previous_row = row_lookup.get((brand, mass_split, prev_period[0], prev_period[1]))
                if not is_estimate_row and has_pulse and k_row_current is not None:
                    ws.cell(row, 12).value = f"=$K${k_row_current}*J{row}"
                elif previous_row is not None:
                    ws.cell(row, 12).value = f"=L{previous_row}/(1+I{row})"
                elif k_row_fallback is not None:
                    ws.cell(row, 12).value = f"=$K${k_row_fallback}*J{row}"
                else:
                    ws.cell(row, 12).value = 0

            else:
                if denominator_in_split:
                    denominator = split_quarter_sum_ref(
                        split_file_name, country, brand, mass_split,
                        denominator_year, denominator_q_start
                    )
                else:
                    rows        = quarter_rows(row_lookup, brand, mass_split, denominator_year, denominator_q_start)
                    denominator = sum_range(rows)
                ws.cell(row, 10).value = f"=H{row}/{denominator}"

                should_hold_actual = not is_estimate_row and has_pulse and month == q_start
                if should_hold_actual:
                    col = report_quarters[(year, quarter)]
                    ws.cell(row, 11).value = external_formula(
                        report_file_name, report_sheet, f"{col}{report_row}"
                    )
                else:
                    ws.cell(row, 11).value = None

                k_row_current  = row_lookup.get((brand, mass_split, year, q_start))
                k_row_fallback = row_lookup.get((brand, mass_split, year - 1, q_start))
                prev_period    = previous_month(year, month)
                previous_row   = row_lookup.get((brand, mass_split, prev_period[0], prev_period[1]))

                if not is_estimate_row and has_pulse and k_row_current is not None:
                    ws.cell(row, 12).value = f"=$K${k_row_current}*J{row}"
                elif previous_row is not None:
                    ws.cell(row, 12).value = f"=L{previous_row}/(1+I{row})"
                elif k_row_fallback is not None:
                    ws.cell(row, 12).value = f"=$K${k_row_fallback}*J{row}"
                else:
                    ws.cell(row, 12).value = 0

            ws.cell(row,  9).number_format = "0.00%"
            ws.cell(row, 10).number_format = "0.00%"
            for col in [8, 11, 12]:
                ws.cell(row, col).number_format = "#,##0.00"


# ===========================================================================
# CLI entry point
# ===========================================================================

def resolve_path(base_dir: Path, value: str | Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else base_dir / path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Add/update one monthly LDB offline estimation workbook (SUNCARE) with live formulas."
    )
    parser.add_argument(
        "target_month",
        nargs="?",
        type=parse_target_month,
        help="Target month in YYYY-MM format, e.g. 2026-04",
    )
    parser.add_argument("--ldb-dir",              type=Path, default=DEFAULT_LDB_DIR)
    parser.add_argument("--input",                default=None)
    parser.add_argument("--output",               default=None)
    parser.add_argument("--split-file",           default=DEFAULT_SPLIT_FILE)
    parser.add_argument("--output-template",      default=DEFAULT_OUTPUT_TEMPLATE)
    parser.add_argument("--final-start-year",     type=int, default=2024)
    parser.add_argument("--final-output",         default=None)
    parser.add_argument("--oo-file",              default=DEFAULT_OO_FILE)
    parser.add_argument("--skip-split-build",     action="store_true")
    parser.add_argument("--export-picked-values", default=None)
    parser.add_argument("--my-report",            default=DEFAULT_MY_REPORT)
    parser.add_argument("--sg-report",            default=DEFAULT_SG_REPORT)
    args = parser.parse_args()

    if args.target_month:
        target_year, target_month = args.target_month
    else:
        while True:
            text = input("Enter target month (YYYY-MM), for example 2026-04: ").strip()
            try:
                target_year, target_month = parse_target_month(text)
                break
            except argparse.ArgumentTypeError as exc:
                print(f"Invalid month: {exc}")

    label      = month_label(target_year, target_month)
    prev_y, prev_m = previous_month(target_year, target_month)
    prev_label = month_label(prev_y, prev_m)

    default_workbook  = f"Offline Estimation {label}.xlsx"
    previous_workbook = f"Offline Estimation {prev_label}.xlsx"

    if args.input:
        input_name = args.input
    elif (args.ldb_dir / default_workbook).exists():
        input_name = default_workbook
    else:
        input_name = previous_workbook

    output_name       = args.output or default_workbook
    final_output_name = args.final_output or f"Suncare LDB Data Output {label}.xlsx"

    ldb_dir           = args.ldb_dir
    input_path        = ldb_dir / input_name
    output_path       = ldb_dir / output_name
    final_output_path = ldb_dir / final_output_name
    my_report_path    = resolve_path(ldb_dir, args.my_report)
    sg_report_path    = resolve_path(ldb_dir, args.sg_report)

    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    required_paths = [my_report_path, sg_report_path]
    if args.skip_split_build:
        required_paths.append(ldb_dir / args.split_file)
    else:
        required_paths.append(ldb_dir / args.oo_file)
    for p in required_paths:
        if not p.exists():
            raise FileNotFoundError(f"Required workbook not found: {p}")

    if not (ldb_dir / args.output_template).exists():
        raise FileNotFoundError(f"Output template not found: {ldb_dir / args.output_template}")

    split_path = ldb_dir / args.split_file
    if args.skip_split_build:
        split_max_year, split_max_month = read_split_max_period(split_path)
    else:
        split_max_year, split_max_month = build_split_method_workbook(
            ldb_dir / args.oo_file, split_path
        )

    if args.export_picked_values:
        export_path = ldb_dir / args.export_picked_values
        row_count   = export_oo_picked_values(ldb_dir / args.oo_file, export_path)
        print(f"Exported {row_count} O+O picked values: {export_path}")

    report_info = {
        "MY": find_report_rows_and_quarters(my_report_path, "MY"),
        "SG": find_report_rows_and_quarters(sg_report_path, "SG"),
    }

    for c in ["MY", "SG"]:
        _, quarters = report_info[c]
        print(f"\n{c} quarters found:")
        print(quarters)

    wb = load_workbook(input_path)
    for sheet_name, country, report_path in [
        ("MY LDB", "MY", my_report_path),
        ("SG LDB", "SG", sg_report_path),
    ]:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {sheet_name}")
        report_rows, report_quarters = report_info[country]
        update_formulas(
            wb[sheet_name],
            target_year,
            target_month,
            split_max_year,
            split_max_month,
            args.split_file,
            report_path.name,
            report_rows,
            report_quarters,
        )

    wb.calculation.calcMode       = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc  = True
    wb.save(output_path)

    sellout_values = calculate_sellout_values(
        wb,
        split_path,
        {"MY": my_report_path, "SG": sg_report_path},
        report_info,
        split_max_year,
        split_max_month,
        target_year,
        target_month,
    )

    my_top = calculate_competitor_values(
        my_report_path,
        split_path,
        "MY",
        target_year,
        target_month,
        report_info["MY"][1],  # report_quarters for MY
    )

    sg_top = calculate_competitor_values(
        sg_report_path,
        split_path,
        "SG",
        target_year,
        target_month,
        report_info["SG"][1],  # report_quarters for SG
    )

    final_rows = export_final_output(
        wb,
        ldb_dir / args.output_template,
        final_output_path,
        args.final_start_year,
        target_year,
        target_month,
        sellout_values,
        my_top,
        sg_top,
    )

    print(f"Updated split method through {split_max_year}-{split_max_month:02d}: {split_path}")
    print(f"Updated {output_path}")
    print(f"Updated final output with {final_rows} rows: {final_output_path}")


if __name__ == "__main__":
    main()
