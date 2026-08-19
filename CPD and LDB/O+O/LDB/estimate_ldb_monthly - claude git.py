from __future__ import annotations
import argparse
import calendar
import csv
import sys
import time
from copy import copy
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font, PatternFill


def log(message: str) -> None:
    """Print progress immediately (unbuffered) so long steps don't look frozen."""
    print(message, flush=True)


class Stopwatch:
    """Small helper to log how long a step took."""

    def __init__(self, label: str):
        self.label = label

    def __enter__(self):
        log(f"-> {self.label} ...")
        self._start = time.perf_counter()
        return self

    def __exit__(self, exc_type, exc, tb):
        elapsed = time.perf_counter() - self._start
        if exc_type is None:
            log(f"   done in {elapsed:.1f}s")
        else:
            log(f"   failed after {elapsed:.1f}s")
DEFAULT_LDB_DIR = Path(__file__).resolve().parent / "LDB"
DEFAULT_REPORT_DIR = Path(__file__).resolve().parent / "Hand and Body Pulse reports"
DEFAULT_SPLIT_FILE = "Bodycare split method and estimation.xlsx"
DEFAULT_OO_FILE = "MYSG ONE LDB CMI YTD Jun'26.xlsx"
DEFAULT_OUTPUT_TEMPLATE = "Data Output format.xlsx"
DEFAULT_MY_REPORT = DEFAULT_REPORT_DIR / "MY L'OREAL_Hand & Body Mositurizer Report - 2026-04-20 (1).xlsx"
DEFAULT_SG_REPORT = DEFAULT_REPORT_DIR / "SG L'OREAL_Hand & Body Mositurizer Report - 2026-04-20 (1).xlsx"
REPORT_SHEET = "1-Total Category & LOreal"
MARKET_REPORT_SHEET = "4-Total Derma & Top 10 brands"
REPORT_HEADER_ROW = 10
REPORT_PRODUCT_COL = 1
BRAND_REPORT_PRODUCTS = {
    "MY": {
        "BRAND_06": (REPORT_SHEET, "CERAVE"),
        "BRAND_07": (REPORT_SHEET, "LA ROCHE POSAY"),
        "Market": (MARKET_REPORT_SHEET, "TOTAL DERMA"),
    },
    "SG": {
        "BRAND_06": (REPORT_SHEET, "CERAVE"),
        "BRAND_07": (REPORT_SHEET, "LA ROCHE POSAY"),
        "Market": (MARKET_REPORT_SHEET, "TOTAL DERMA"),
    },
}
SPLIT_SOURCE_ROWS = {
    ("MY", "BRAND_06", None): ("Brand", 3),
    ("MY", "BRAND_07", None): ("Brand", 4),
    ("SG", "BRAND_06", None): ("Brand", 5),
    ("SG", "BRAND_07", None): ("Brand", 6),
    ("MY", "Market", "Mass Medical"): ("Market", 3),
    ("MY", "Market", "Non-Mass Medical"): ("Market", 4),
    ("SG", "Market", "Mass Medical"): ("Market", 5),
    ("SG", "Market", "Non-Mass Medical"): ("Market", 6),
}
BRAND_SOURCE_MAP = {
    "CERAVE": "BRAND_06",
    "LA ROCHE POSAY": "BRAND_07",
}
OUTPUT_BRAND_MAP = {
    ("Market", "Mass Medical"): ("Medic Market", "Mass Medical"),
    ("Market", "Non-Mass Medical"): ("Medic Market", "Non-Mass Medical"),
    ("BRAND_06", None): ("Cerave", "Mass Medical"),
    ("BRAND_07", None): ("La Roche Posay", "Non-Mass Medical"),
}
FINAL_OUTPUT_ORDER = [
    ("Market", "Mass Medical"),
    ("Market", "Non-Mass Medical"),
    ("BRAND_06", None),
    ("BRAND_07", None),
]
def parse_target_month(value: str) -> tuple[int, int]:
    try:
        year_text, month_text = value.split("-", 1)
        year = int(year_text)
        month = int(month_text)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Use YYYY-MM, for example 2026-04") from exc
    if month < 1 or month > 12:
        raise argparse.ArgumentTypeError("Month must be between 01 and 12")
    return year, month
def month_label(year: int, month: int) -> str:
    return f"{calendar.month_abbr[month]}'{str(year)[-2:]}"
def quote_sheet(name: str) -> str:
    return name.replace("'", "''")
def external_ref(workbook_name: str, sheet_name: str, cell_ref: str) -> str:
    workbook_name = workbook_name.replace("'", "''")
    return f"'[{workbook_name}]{quote_sheet(sheet_name)}'!${cell_ref}"
def external_formula(workbook_name: str, sheet_name: str, cell_ref: str) -> str:
    return f"={external_ref(workbook_name, sheet_name, cell_ref)}"
def excel_col(number: int) -> str:
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result
def split_month_col(sheet_type: str, year: int, month: int) -> str:
    start_col = 4 if sheet_type == "Brand" else 5
    return excel_col(start_col + (year - 2024) * 12 + (month - 1))
def split_source_cell(country: str, brand: str, mass_split: str | None, year: int, month: int) -> tuple[str, str]:
    split_sheet, split_row = SPLIT_SOURCE_ROWS[(country, brand, mass_split)]
    split_col = split_month_col(split_sheet, year, month)
    return split_sheet, f"{split_col}{split_row}"
def split_source_ref(
    split_file_name: str,
    country: str,
    brand: str,
    mass_split: str | None,
    year: int,
    month: int,
) -> str:
    split_sheet, cell_ref = split_source_cell(country, brand, mass_split, year, month)
    return external_ref(split_file_name, split_sheet, cell_ref)
def split_quarter_sum_ref(
    split_file_name: str,
    country: str,
    brand: str,
    mass_split: str | None,
    year: int,
    q_start: int,
) -> str:
    split_sheet, first_cell = split_source_cell(country, brand, mass_split, year, q_start)
    _, last_cell = split_source_cell(country, brand, mass_split, year, q_start + 2)
    workbook_name = split_file_name.replace("'", "''")
    prefix = f"'[{workbook_name}]{quote_sheet(split_sheet)}'!"
    return f"SUM({prefix}${first_cell}:${last_cell})"
def period_key(year: int, month: int) -> int:
    return year * 12 + month
def quarter_start(month: int) -> int:
    return ((month - 1) // 3) * 3 + 1
def previous_month(year: int, month: int) -> tuple[int, int]:
    if month == 1:
        return year - 1, 12
    return year, month - 1
def capped_actual_period(
    split_max_year: int,
    split_max_month: int,
    target_year: int,
    target_month: int,
) -> tuple[int, int]:
    previous_year, previous_month_number = previous_month(target_year, target_month)
    if period_key(split_max_year, split_max_month) <= period_key(previous_year, previous_month_number):
        return split_max_year, split_max_month
    return previous_year, previous_month_number
def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.number_format:
            target.number_format = source.number_format
def get_groups(ws) -> list[tuple[int, int, str, str | None]]:
    groups = []
    start = 2
    for row in range(3, ws.max_row + 2):
        same_group = False
        if row <= ws.max_row:
            same_group = (
                ws.cell(row, 5).value == ws.cell(row - 1, 5).value
                and ws.cell(row, 7).value == ws.cell(row - 1, 7).value
            )
        if not same_group:
            groups.append((start, row - 1, ws.cell(start, 5).value, ws.cell(start, 7).value))
            start = row
    return groups
def build_row_lookup(ws) -> dict[tuple[str, str | None, int, int], int]:
    lookup = {}
    for row in range(2, ws.max_row + 1):
        brand = ws.cell(row, 5).value
        mass_split = ws.cell(row, 7).value
        year = ws.cell(row, 3).value
        month = ws.cell(row, 4).value
        if brand and year and month:
            lookup[(brand, mass_split, int(year), int(month))] = row
    return lookup
def load_readonly_workbook(path: Path):
    return load_workbook(path, data_only=True, read_only=True)


def find_report_rows_and_quarters(wb, report_name: str, country: str) -> tuple[dict[str, tuple[str, int]], dict[tuple[int, int], str]]:
    row_by_brand = {}
    targets = BRAND_REPORT_PRODUCTS[country]
    for brand, (sheet_name, expected) in targets.items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"{report_name}: missing report sheet {sheet_name!r}")
        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            product = ws.cell(row, REPORT_PRODUCT_COL).value
            if not product:
                continue
            product_text = str(product).strip().upper()
            if product_text == expected.upper():
                row_by_brand[brand] = (sheet_name, row)
                break
    col_by_quarter = {}
    ws = wb[REPORT_SHEET]
    for col in range(2, ws.max_column + 1):
        if str(ws.cell(REPORT_HEADER_ROW - 1, col).value or "").strip() != "Sales Value":
            continue
        header = ws.cell(REPORT_HEADER_ROW, col).value
        if not header:
            continue
        text = str(header)
        if not text.startswith("Q"):
            continue
        # Example: Q1 26 - w/e 31/03/26
        quarter = int(text[1])
        year = 2000 + int(text.split()[1])
        col_by_quarter[(year, quarter)] = excel_col(col)
    missing = sorted(set(targets) - set(row_by_brand))
    if missing:
        raise ValueError(f"{report_name}: missing report rows for {', '.join(missing)}")
    return row_by_brand, col_by_quarter
def find_country_blocks(ws) -> dict[str, int]:
    blocks = {}
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value != "Country (Currency)":
            continue
        country_text = str(ws.cell(row, 2).value or "").upper()
        if country_text.startswith("MY "):
            blocks["MY"] = row
        elif country_text.startswith("SG "):
            blocks["SG"] = row
    missing = sorted({"MY", "SG"} - set(blocks))
    if missing:
        raise ValueError(f"{ws.title}: missing country block(s): {', '.join(missing)}")
    return blocks
def extract_period_columns(ws, year_row: int, month_row: int, start_col: int) -> list[tuple[int, int, int]]:
    periods = []
    for col in range(start_col, ws.max_column + 1):
        year = ws.cell(year_row, col).value
        month = ws.cell(month_row, col).value
        if isinstance(year, int) and isinstance(month, int):
            periods.append((col, year, month))
    if not periods:
        raise ValueError(f"{ws.title}: no monthly period columns found")
    return periods
def copy_period_headers(ws, start_col: int, periods: list[tuple[int, int, int]]) -> None:
    for offset, (_, year, month) in enumerate(periods):
        col = start_col + offset
        ws.cell(1, col).value = year
        ws.cell(2, col).value = month
def find_last_split_value_period(wb) -> tuple[int, int]:
    periods = []
    for sheet_name, start_col, data_row_start in [("Brand", 4, 3), ("Market", 5, 3)]:
        ws = wb[sheet_name]
        for col in range(start_col, ws.max_column + 1):
            year = ws.cell(1, col).value
            month = ws.cell(2, col).value
            if not (isinstance(year, int) and isinstance(month, int)):
                continue
            has_value = any(ws.cell(row, col).value is not None for row in range(data_row_start, ws.max_row + 1))
            if has_value:
                periods.append((year, month))
    if not periods:
        raise ValueError("Split workbook has no monthly values")
    return max(periods)
def build_split_method_workbook(oo_path: Path, output_path: Path) -> tuple[int, int]:
    source_wb = load_workbook(oo_path, data_only=True, read_only=True)
    for sheet_name in ["Market", "L'Oreal "]:
        if sheet_name not in source_wb.sheetnames:
            raise ValueError(f"{oo_path.name}: missing sheet {sheet_name!r}")
    market_source = source_wb["Market"]
    brand_source = source_wb["L'Oreal "]
    market_blocks = find_country_blocks(market_source)
    brand_blocks = find_country_blocks(brand_source)
    market_periods = extract_period_columns(market_source, market_blocks["MY"] + 4, market_blocks["MY"] + 5, 4)
    brand_periods = extract_period_columns(brand_source, brand_blocks["MY"] + 4, brand_blocks["MY"] + 5, 3)
    output_wb = Workbook()
    market_ws = output_wb.active
    market_ws.title = "Market"
    brand_ws = output_wb.create_sheet("Brand")
    market_headers = ["Country", "O+O Brand", "Mass/Non-Mass", "Category"]
    brand_headers = ["Country", "O+O Brand", "Category"]
    for col, value in enumerate(market_headers, start=1):
        market_ws.cell(2, col).value = value
    for col, value in enumerate(brand_headers, start=1):
        brand_ws.cell(2, col).value = value
    copy_period_headers(market_ws, 5, market_periods)
    copy_period_headers(brand_ws, 4, brand_periods)
    market_output_row = 3
    for country in ["MY", "SG"]:
        block = market_blocks[country]
        for source_row in [block + 6, block + 7]:
            market_ws.cell(market_output_row, 1).value = country
            market_ws.cell(market_output_row, 2).value = market_source.cell(source_row, 1).value
            market_ws.cell(market_output_row, 3).value = market_source.cell(source_row, 2).value
            market_ws.cell(market_output_row, 4).value = market_source.cell(source_row, 3).value
            for offset, (source_col, _, _) in enumerate(market_periods):
                market_ws.cell(market_output_row, 5 + offset).value = market_source.cell(source_row, source_col).value
            market_output_row += 1
    brand_output_row = 3
    for country in ["MY", "SG"]:
        block = brand_blocks[country]
        for source_row in [block + 6, block + 7]:
            source_brand = str(brand_source.cell(source_row, 1).value or "").strip().upper()
            brand_ws.cell(brand_output_row, 1).value = country
            brand_ws.cell(brand_output_row, 2).value = BRAND_SOURCE_MAP.get(source_brand, source_brand)
            brand_ws.cell(brand_output_row, 3).value = brand_source.cell(source_row, 2).value
            for offset, (source_col, _, _) in enumerate(brand_periods):
                brand_ws.cell(brand_output_row, 4 + offset).value = brand_source.cell(source_row, source_col).value
            brand_output_row += 1
    for ws in [market_ws, brand_ws]:
        header_fill = PatternFill("solid", fgColor="D9EAD3")
        for row in [1, 2]:
            for cell in ws[row]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
        for column_cells in ws.columns:
            ws.column_dimensions[column_cells[0].column_letter].width = 14
    output_wb.save(output_path)
    return find_last_split_value_period(output_wb)
def export_oo_picked_values(oo_path: Path, output_path: Path) -> int:
    source_wb = load_workbook(oo_path, data_only=True, read_only=True)
    rows = []
    brand_source = source_wb["L'Oreal "]
    brand_blocks = find_country_blocks(brand_source)
    for country in ["MY", "SG"]:
        block = brand_blocks[country]
        periods = extract_period_columns(brand_source, block + 4, block + 5, 3)
        for source_row in [block + 6, block + 7]:
            raw_brand = str(brand_source.cell(source_row, 1).value or "").strip()
            output_brand = BRAND_SOURCE_MAP.get(raw_brand.upper(), raw_brand)
            category = brand_source.cell(source_row, 2).value
            for source_col, year, month in periods:
                rows.append(
                    {
                        "Output Sheet": "Brand",
                        "Country": country,
                        "Source Sheet": "L'Oreal ",
                        "Source Cell": brand_source.cell(source_row, source_col).coordinate,
                        "Source Brand": raw_brand,
                        "Output Brand": output_brand,
                        "Mass/Non-Mass": "",
                        "Category": category,
                        "Year": year,
                        "Month": month,
                        "Value": brand_source.cell(source_row, source_col).value,
                    }
                )
    market_source = source_wb["Market"]
    market_blocks = find_country_blocks(market_source)
    for country in ["MY", "SG"]:
        block = market_blocks[country]
        periods = extract_period_columns(market_source, block + 4, block + 5, 4)
        for source_row in [block + 6, block + 7]:
            brand = market_source.cell(source_row, 1).value
            mass_split = market_source.cell(source_row, 2).value
            category = market_source.cell(source_row, 3).value
            for source_col, year, month in periods:
                rows.append(
                    {
                        "Output Sheet": "Market",
                        "Country": country,
                        "Source Sheet": "Market",
                        "Source Cell": market_source.cell(source_row, source_col).coordinate,
                        "Source Brand": brand,
                        "Output Brand": brand,
                        "Mass/Non-Mass": mass_split,
                        "Category": category,
                        "Year": year,
                        "Month": month,
                        "Value": market_source.cell(source_row, source_col).value,
                    }
                )
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)
def estimation_external_formula(workbook_name: str, sheet_name: str, cell_ref: str) -> str:
    return f"={external_ref(workbook_name, sheet_name, cell_ref)}"
def read_split_value(split_wb, country: str, brand: str, mass_split: str | None, year: int, month: int) -> float:
    sheet_name, cell_ref = split_source_cell(country, brand, mass_split, year, month)
    value = split_wb[sheet_name][cell_ref].value
    if value is None:
        raise ValueError(f"Missing split value for {country} {brand} {mass_split or ''} {year}-{month:02d}")
    return float(value)
def read_report_value(report_wb, report_rows: dict[str, tuple[str, int]], report_quarters: dict[tuple[int, int], str], brand: str, year: int, quarter: int) -> float:
    sheet_name, row = report_rows[brand]
    col_letter = report_quarters[(year, quarter)]
    value = report_wb[sheet_name].cell(row, column_index_from_string(col_letter)).value
    if value is None:
        raise ValueError(f"Missing pulse report value for {brand} {year} Q{quarter}")
    return float(value)
def calculate_sellout_values(
    estimation_wb,
    split_path: Path,
    report_wbs: dict[str, object],
    report_info: dict[str, tuple[dict[str, int], dict[tuple[int, int], str]]],
    split_max_year: int,
    split_max_month: int,
    target_year: int,
    target_month: int,
) -> dict[tuple[str, int], float]:
    split_wb = load_workbook(split_path, data_only=True, read_only=True)
    actual_max_year, actual_max_month = capped_actual_period(
        split_max_year, split_max_month, target_year, target_month
    )
    split_max_key = period_key(actual_max_year, actual_max_month)
    calculated = {}
    for sheet_name in ["MY LDB", "SG LDB"]:
        ws = estimation_wb[sheet_name]
        country = sheet_name[:2]
        report_rows, report_quarters = report_info[country]
        row_lookup = build_row_lookup(ws)
        groups = get_groups(ws)
        h_values = {}
        i_values = {}
        j_values = {}
        k_values = {}
        l_values = {}
        for start, end, brand, mass_split in groups:
            for row in range(start, end + 1):
                year = int(ws.cell(row, 3).value)
                month = int(ws.cell(row, 4).value)
                q_start = quarter_start(month)
                quarter = (month - 1) // 3 + 1
                has_pulse = (year, quarter) in report_quarters
                is_estimate_row = period_key(year, month) > split_max_key or not has_pulse
                if is_estimate_row:
                    # 2024 is the starting year, so do not look for 2023 data
                    if year == 2024:
                        h_values[row] = read_split_value(split_wb, country, brand, mass_split, year, month)
                    else:
                        prev_year, prev_month = previous_month(year, month)
                        previous_row = row_lookup[(brand, mass_split, prev_year, prev_month)]
                        prior_year_row = row_lookup[(brand, mass_split, year - 1, month)]
                        h_values[row] = h_values[previous_row] / (1 + i_values[prior_year_row])
                else:
                    h_values[row] = read_split_value(split_wb, country, brand, mass_split, year, month)
                if row == start:
                    i_values[row] = None
                else:
                    i_values[row] = h_values[row - 1] / h_values[row] - 1
                denominator_year = year - 1 if is_estimate_row else year
                denominator_q_start = q_start
                denominator_in_split = period_key(denominator_year, denominator_q_start + 2) <= split_max_key
                if brand == "Market":
                    if not is_estimate_row or denominator_in_split:
                        denominator = sum(
                            read_split_value(split_wb, country, "Market", split, denominator_year, q_month)
                            for split in ["Mass Medical", "Non-Mass Medical"]
                            for q_month in range(denominator_q_start, denominator_q_start + 3)
                        )
                    else:
                        denominator = sum(
                            h_values[row_lookup[("Market", split, denominator_year, q_month)]]
                            for split in ["Mass Medical", "Non-Mass Medical"]
                            for q_month in range(denominator_q_start, denominator_q_start + 3)
                        )
                    j_values[row] = h_values[row] / denominator
                    should_hold_actual = (
                        not is_estimate_row
                        and has_pulse
                        and mass_split == "Mass Medical"
                        and month == q_start
                    )
                    if should_hold_actual:
                        k_values[row] = read_report_value(
                            report_wbs[country], report_rows, report_quarters, brand, year, quarter
                        )
                    else:
                        k_values[row] = None
                    k_row_current = row_lookup.get(("Market", "Mass Medical", year, q_start))
                    previous_period = previous_month(year, month)
                    previous_row = row_lookup.get((brand, mass_split, previous_period[0], previous_period[1]))
                    if not is_estimate_row and has_pulse and k_row_current:
                        l_values[row] = k_values[k_row_current] * j_values[row]
                    elif previous_row:
                        l_values[row] = l_values[previous_row] / (1 + i_values[row])
                    else:
                        k_row_fallback = row_lookup[("Market", "Mass Medical", year - 1, q_start)]
                        l_values[row] = k_values[k_row_fallback] * j_values[row]
                else:
                    if not is_estimate_row or denominator_in_split:
                        denominator = sum(
                            read_split_value(split_wb, country, brand, mass_split, denominator_year, q_month)
                            for q_month in range(denominator_q_start, denominator_q_start + 3)
                        )
                    else:
                        denominator = sum(
                            h_values[row_lookup[(brand, mass_split, denominator_year, q_month)]]
                            for q_month in range(denominator_q_start, denominator_q_start + 3)
                        )
                    j_values[row] = h_values[row] / denominator
                    should_hold_actual = not is_estimate_row and has_pulse and month == q_start
                    if should_hold_actual:
                        k_values[row] = read_report_value(
                            report_wbs[country], report_rows, report_quarters, brand, year, quarter
                        )
                    else:
                        k_values[row] = None
                    k_row_current = row_lookup.get((brand, mass_split, year, q_start))
                    previous_period = previous_month(year, month)
                    previous_row = row_lookup.get((brand, mass_split, previous_period[0], previous_period[1]))
                    if not is_estimate_row and has_pulse and k_row_current:
                        l_values[row] = k_values[k_row_current] * j_values[row]
                    elif previous_row:
                        l_values[row] = l_values[previous_row] / (1 + i_values[row])
                    else:
                        k_row_fallback = row_lookup[(brand, mass_split, year - 1, q_start)]
                        l_values[row] = k_values[k_row_fallback] * j_values[row]
        for row, value in l_values.items():
            calculated[(sheet_name, row)] = value
    return calculated
def export_final_output(
    estimation_wb,
    template_path: Path,
    output_path: Path,
    start_year: int,
    target_year: int,
    target_month: int,
    sellout_values: dict[tuple[str, int], float],
) -> int:
    if not template_path.exists():
        raise FileNotFoundError(f"Output template not found: {template_path}")
    output_wb = load_workbook(template_path)
    template_ws = output_wb.active
    header_values = [template_ws.cell(1, col).value for col in range(1, template_ws.max_column + 1)]
    header_styles = []
    for col in range(1, template_ws.max_column + 1):
        source_cell = template_ws.cell(1, col)
        header_styles.append(
            {
                "style": copy(source_cell._style),
                "font": copy(source_cell.font),
                "fill": copy(source_cell.fill),
                "border": copy(source_cell.border),
                "alignment": copy(source_cell.alignment),
                "number_format": source_cell.number_format,
                "width": template_ws.column_dimensions[source_cell.column_letter].width,
            }
        )
    for sheet_name in list(output_wb.sheetnames):
        del output_wb[sheet_name]
    output_sheets = {}
    for country in ["MY", "SG"]:
        ws = output_wb.create_sheet(country)
        output_sheets[country] = ws
        for col, value in enumerate(header_values, start=1):
            cell = ws.cell(1, col)
            cell.value = value
            style = header_styles[col - 1]
            cell._style = copy(style["style"])
            cell.font = copy(style["font"])
            cell.fill = copy(style["fill"])
            cell.border = copy(style["border"])
            cell.alignment = copy(style["alignment"])
            cell.number_format = style["number_format"]
            ws.column_dimensions[cell.column_letter].width = style["width"]
    row_lookups = {
        sheet_name: build_row_lookup(estimation_wb[sheet_name])
        for sheet_name in ["MY LDB", "SG LDB"]
    }
    output_rows = {"MY": 2, "SG": 2}
    for year in range(start_year, target_year + 1):
        month_start = 1
        month_end = target_month if year == target_year else 12
        for month in range(month_start, month_end + 1):
            for sheet_name in ["MY LDB", "SG LDB"]:
                country = sheet_name[:2]
                ws = output_sheets[country]
                output_row = output_rows[country]
                row_lookup = row_lookups[sheet_name]
                for brand, mass_split in FINAL_OUTPUT_ORDER:
                    source_row = row_lookup.get((brand, mass_split, year, month))
                    if source_row is None:
                        raise ValueError(
                            f"{sheet_name}: missing final output source row for "
                            f"{brand} {mass_split or ''} {year}-{month:02d}"
                        )
                    output_brand, output_mass_split = OUTPUT_BRAND_MAP[(brand, mass_split)]
                    ws.cell(output_row, 1).value = country
                    ws.cell(output_row, 2).value = "Offline"
                    ws.cell(output_row, 3).value = year
                    ws.cell(output_row, 4).value = "Offline_Est"
                    ws.cell(output_row, 5).value = "SKINCARE"
                    ws.cell(output_row, 6).value = "Body Care"
                    ws.cell(output_row, 7).value = output_mass_split
                    ws.cell(output_row, 8).value = output_brand
                    ws.cell(output_row, 9).value = month
                    ws.cell(output_row, 10).value = sellout_values[(sheet_name, source_row)]
                    ws.cell(output_row, 10).number_format = "#,##0.00"
                    ws.cell(output_row, 11).value = None
                    output_row += 1
                output_rows[country] = output_row
    output_wb.save(output_path)
    return sum(row - 2 for row in output_rows.values())
def read_split_max_period(split_path: Path) -> tuple[int, int]:
    wb = load_workbook(split_path, data_only=True, read_only=True)
    try:
        return find_last_split_value_period(wb)
    except ValueError as exc:
        raise ValueError(f"{split_path.name}: {exc}") from exc
def resolve_path(base_dir: Path, value: str | Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else base_dir / path
def ensure_target_rows(ws, target_year: int, target_month: int) -> None:
    groups = get_groups(ws)
    row_lookup = build_row_lookup(ws)
    inserts = []
    for start, end, brand, mass_split in groups:
        if (brand, mass_split, target_year, target_month) in row_lookup:
            continue
        prev_year, prev_month = previous_month(target_year, target_month)
        previous_row = row_lookup.get((brand, mass_split, prev_year, prev_month))
        if previous_row is None:
            raise ValueError(
                f"{ws.title}: cannot add {target_year}-{target_month:02d} for "
                f"{brand} {mass_split or ''}; previous month is missing"
            )
        if previous_row != end:
            raise ValueError(
                f"{ws.title}: {brand} {mass_split or ''} has rows after the previous month. "
                "Run missing months in order."
            )
        inserts.append(previous_row)
    for previous_row in sorted(inserts, reverse=True):
        ws.insert_rows(previous_row + 1, 1)
        copy_row_style(ws, previous_row, previous_row + 1)
        for col in [1, 2, 5, 6, 7, 13]:
            ws.cell(previous_row + 1, col).value = ws.cell(previous_row, col).value
        ws.cell(previous_row + 1, 3).value = target_year
        ws.cell(previous_row + 1, 4).value = target_month
def quarter_rows(row_lookup, brand, mass_split, year, q_start) -> list[int]:
    return [
        row_lookup[(brand, mass_split, year, month)]
        for month in range(q_start, q_start + 3)
        if (brand, mass_split, year, month) in row_lookup
    ]
def sum_range(rows: list[int]) -> str:
    if not rows:
        return "0"
    return f"SUM(H{min(rows)}:H{max(rows)})"
def update_formulas(
    ws,
    target_year: int,
    target_month: int,
    split_max_year: int,
    split_max_month: int,
    split_file_name: str,
    report_file_name: str,
    report_rows: dict[str, int],
    report_quarters: dict[tuple[int, int], str],
) -> None:
    country = ws.title[:2]
    ensure_target_rows(ws, target_year, target_month)
    row_lookup = build_row_lookup(ws)
    groups = get_groups(ws)
    actual_max_year, actual_max_month = capped_actual_period(
        split_max_year, split_max_month, target_year, target_month
    )
    split_max_key = period_key(actual_max_year, actual_max_month)
    for start, end, brand, mass_split in groups:
        for row in range(start, end + 1):
            year = int(ws.cell(row, 3).value)
            month = int(ws.cell(row, 4).value)
            q_start = quarter_start(month)
            quarter = (month - 1) // 3 + 1
            has_pulse = (year, quarter) in report_quarters
            is_estimate_row = period_key(year, month) > split_max_key or not has_pulse
            report_sheet, report_row = report_rows[brand]
            ws.cell(row, 13).value = report_sheet
            if is_estimate_row:
                # 2024 is the starting year, so do not look for 2023 data
                if year == 2024:
                    # Keep the existing 2024 split/source logic
                    split_sheet, cell_ref = split_source_cell(
                        country, brand, mass_split, year, month
                    )
                    ws.cell(row, 8).value = external_formula(
                        split_file_name, split_sheet, cell_ref
                    )
                else:
                    prev_year, prev_month = previous_month(year, month)
                    previous_row = row_lookup[(brand, mass_split, prev_year, prev_month)]
                    prior_year_row = row_lookup[(brand, mass_split, year - 1, month)]
                    ws.cell(row, 8).value = f"=H{previous_row}/(1+I{prior_year_row})"
            else:
                split_sheet, cell_ref = split_source_cell(country, brand, mass_split, year, month)
                ws.cell(row, 8).value = external_formula(split_file_name, split_sheet, cell_ref)
            ws.cell(row, 9).value = None if row == start else f"=H{row - 1}/H{row}-1"
            denominator_year = year - 1 if is_estimate_row else year
            denominator_q_start = q_start
            if brand == "Market":
                if not is_estimate_row:
                    mass_denominator = split_quarter_sum_ref(
                        split_file_name, country, "Market", "Mass Medical", denominator_year, denominator_q_start
                    )
                    non_mass_denominator = split_quarter_sum_ref(
                        split_file_name, country, "Market", "Non-Mass Medical", denominator_year, denominator_q_start
                    )
                    denominator = f"{mass_denominator}+{non_mass_denominator}"
                else:
                    mass_rows = quarter_rows(
                        row_lookup, "Market", "Mass Medical", denominator_year, denominator_q_start
                    )
                    non_mass_rows = quarter_rows(
                        row_lookup, "Market", "Non-Mass Medical", denominator_year, denominator_q_start
                    )
                    denominator = f"{sum_range(mass_rows)}+{sum_range(non_mass_rows)}"
                ws.cell(row, 10).value = f"=H{row}/({denominator})"
                k_row_current = row_lookup.get(("Market", "Mass Medical", year, q_start))
                k_row_fallback = row_lookup.get(("Market", "Mass Medical", year - 1, q_start))
                should_hold_actual = (
                    not is_estimate_row
                    and
                    has_pulse
                    and
                    mass_split == "Mass Medical"
                    and month == q_start
                )
                if should_hold_actual:
                    col = report_quarters[(year, quarter)]
                    ws.cell(row, 11).value = external_formula(report_file_name, report_sheet, f"{col}{report_row}")
                else:
                    ws.cell(row, 11).value = None
                previous_period = previous_month(year, month)
                previous_row = row_lookup.get((brand, mass_split, previous_period[0], previous_period[1]))
                if not is_estimate_row and has_pulse and k_row_current:
                    ws.cell(row, 12).value = f"=$K${k_row_current}*J{row}"
                elif previous_row:
                    ws.cell(row, 12).value = f"=L{previous_row}/(1+I{row})"
                else:
                    ws.cell(row, 12).value = f"=$K${k_row_fallback}*J{row}"
            else:
                if not is_estimate_row:
                    denominator = split_quarter_sum_ref(
                        split_file_name, country, brand, mass_split, denominator_year, denominator_q_start
                    )
                else:
                    rows = quarter_rows(row_lookup, brand, mass_split, denominator_year, denominator_q_start)
                    denominator = sum_range(rows)
                ws.cell(row, 10).value = f"=H{row}/{denominator}"
                should_hold_actual = not is_estimate_row and has_pulse and month == q_start
                if should_hold_actual:
                    col = report_quarters[(year, quarter)]
                    ws.cell(row, 11).value = external_formula(report_file_name, report_sheet, f"{col}{report_row}")
                else:
                    ws.cell(row, 11).value = None
                k_row_current = row_lookup.get((brand, mass_split, year, q_start))
                k_row_fallback = row_lookup.get((brand, mass_split, year - 1, q_start))
                previous_period = previous_month(year, month)
                previous_row = row_lookup.get((brand, mass_split, previous_period[0], previous_period[1]))
                if not is_estimate_row and has_pulse and k_row_current:
                    ws.cell(row, 12).value = f"=$K${k_row_current}*J{row}"
                elif previous_row:
                    ws.cell(row, 12).value = f"=L{previous_row}/(1+I{row})"
                else:
                    ws.cell(row, 12).value = f"=$K${k_row_fallback}*J{row}"
            ws.cell(row, 9).number_format = "0.00%"
            ws.cell(row, 10).number_format = "0.00%"
            for col in [8, 11, 12]:
                ws.cell(row, col).number_format = "#,##0.00"
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Add/update one monthly LDB offline estimation workbook with live formulas."
    )
    parser.add_argument(
        "target_month",
        nargs="?",
        type=parse_target_month,
        help="Target month in YYYY-MM format, e.g. 2026-04",
    )
    parser.add_argument("--ldb-dir", type=Path, default=DEFAULT_LDB_DIR)
    parser.add_argument("--input", default=None, help="Input workbook name. Defaults to Offline Estimation <Mon>'<YY>.xlsx")
    parser.add_argument("--output", default=None, help="Output workbook name. Defaults to updating the input workbook")
    parser.add_argument("--split-file", default=DEFAULT_SPLIT_FILE)
    parser.add_argument("--output-template", default=DEFAULT_OUTPUT_TEMPLATE)
    parser.add_argument(
        "--final-start-year",
        type=int,
        default=2024,
        help="First year to include in the final flat output. Defaults to 2024.",
    )
    parser.add_argument(
        "--final-output",
        default=None,
        help="Final flat output workbook name. Defaults to LDB Data Output <Mon>'<YY>.xlsx",
    )
    parser.add_argument(
        "--oo-file",
        default=DEFAULT_OO_FILE,
        help="O+O source workbook used to rebuild the split-method workbook",
    )
    parser.add_argument(
        "--skip-split-build",
        action="store_true",
        help="Use the existing split-method workbook instead of rebuilding it from --oo-file",
    )
    parser.add_argument(
        "--export-picked-values",
        default=None,
        help="Optional CSV filename to export the exact O+O values picked for the split-method workbook",
    )
    parser.add_argument("--my-report", default=DEFAULT_MY_REPORT)
    parser.add_argument("--sg-report", default=DEFAULT_SG_REPORT)
    args = parser.parse_args()
    if args.target_month:
        target_year, target_month = args.target_month
    else:
        while True:
            target_month_text = input("Enter target month (YYYY-MM), for example 2026-04: ").strip()
            try:
                target_year, target_month = parse_target_month(target_month_text)
                break
            except argparse.ArgumentTypeError as exc:
                print(f"Invalid month: {exc}")
    default_workbook = f"Offline Estimation {month_label(target_year, target_month)}.xlsx"
    previous_year, previous_month_number = previous_month(target_year, target_month)
    previous_workbook = f"Offline Estimation {month_label(previous_year, previous_month_number)}.xlsx"
    if args.input:
        input_name = args.input
    elif (args.ldb_dir / default_workbook).exists():
        input_name = default_workbook
    else:
        input_name = previous_workbook
    output_name = args.output or default_workbook
    final_output_name = args.final_output or f"LDB Data Output {month_label(target_year, target_month)}.xlsx"
    ldb_dir = args.ldb_dir
    input_path = ldb_dir / input_name
    output_path = ldb_dir / output_name
    final_output_path = ldb_dir / final_output_name
    my_report_path = resolve_path(ldb_dir, args.my_report)
    sg_report_path = resolve_path(ldb_dir, args.sg_report)
    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")
    required_paths = [my_report_path, sg_report_path]
    if args.skip_split_build:
        required_paths.append(ldb_dir / args.split_file)
    else:
        required_paths.append(ldb_dir / args.oo_file)
    for required_path in required_paths:
        if not required_path.exists():
            raise FileNotFoundError(f"Required workbook not found: {required_path}")
    if not (ldb_dir / args.output_template).exists():
        raise FileNotFoundError(f"Output template not found: {ldb_dir / args.output_template}")
    split_path = ldb_dir / args.split_file
    if args.skip_split_build:
        with Stopwatch(f"Reading existing split file max period ({split_path.name})"):
            split_max_year, split_max_month = read_split_max_period(split_path)
    else:
        with Stopwatch(f"Rebuilding split-method workbook from {args.oo_file}"):
            split_max_year, split_max_month = build_split_method_workbook(ldb_dir / args.oo_file, split_path)
    if args.export_picked_values:
        export_path = ldb_dir / args.export_picked_values
        with Stopwatch(f"Exporting O+O picked values to {export_path.name}"):
            row_count = export_oo_picked_values(ldb_dir / args.oo_file, export_path)
        log(f"Exported {row_count} O+O picked values: {export_path}")

    with Stopwatch("Loading pulse report workbooks (MY, SG)"):
        my_report_wb = load_readonly_workbook(my_report_path)
        sg_report_wb = load_readonly_workbook(sg_report_path)

    with Stopwatch("Locating report rows/quarters"):
        report_info = {
            "MY": find_report_rows_and_quarters(my_report_wb, my_report_path.name, "MY"),
            "SG": find_report_rows_and_quarters(sg_report_wb, sg_report_path.name, "SG"),
        }
    report_wbs = {"MY": my_report_wb, "SG": sg_report_wb}

    with Stopwatch(f"Loading estimation workbook ({input_path.name})"):
        wb = load_workbook(input_path)

    for sheet_name, country, report_path in [
        ("MY LDB", "MY", my_report_path),
        ("SG LDB", "SG", sg_report_path),
    ]:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {sheet_name}")
        report_rows, report_quarters = report_info[country]
        row_count = wb[sheet_name].max_row
        with Stopwatch(f"Updating formulas on {sheet_name} ({row_count} rows)"):
            update_formulas(
                wb[sheet_name],
                target_year,
                target_month,
                split_max_year,
                split_max_month,
                args.split_file,
                report_path.name,
                report_rows,
                report_quarters,
            )
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    with Stopwatch(f"Saving {output_path.name}"):
        wb.save(output_path)

    with Stopwatch("Calculating sellout values"):
        sellout_values = calculate_sellout_values(
            wb,
            split_path,
            report_wbs,
            report_info,
            split_max_year,
            split_max_month,
            target_year,
            target_month,
        )
    with Stopwatch(f"Writing final output ({final_output_path.name})"):
        final_rows = export_final_output(
            wb,
            ldb_dir / args.output_template,
            final_output_path,
            args.final_start_year,
            target_year,
            target_month,
            sellout_values,
        )
    log(f"Updated split method through {split_max_year}-{split_max_month:02d}: {split_path}")
    log(f"Updated {output_path}")
    log(f"Updated final output with {final_rows} rows: {final_output_path}")
if __name__ == "__main__":
    main()